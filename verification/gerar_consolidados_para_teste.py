"""
gerar_consolidados_para_teste.py

Gera os arquivos "Consolidado Verification" no formato do template (28 colunas)
a partir dos arquivos "RELATÓRIO CONSOLIDADO" da pasta PARA_TESTE.

Mapeamento de colunas do template:
  1  Veículo          4  Contratado    5  Impressões    7  Cliques
  9  Views            11 Viewables     12 Viewability   13 Entregas Válidas
  14 Acidente*        18 Safeframe     21 Não Classif.  25 Data Ini  26 Data Fim
  27 Devolutiva BI SECOM

* Impróprio (agregado) mapeado para col 14 — fonte não possui breakdown individual.

Uso:
    python3 gerar_consolidados_para_teste.py
"""

import re
import shutil
from copy import copy
from pathlib import Path

import pandas as pd
import openpyxl


# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent
TEMPLATE   = ROOT / "TEMPLATE - Consolidado Verification.xlsx"
PARA_TESTE = ROOT / "Novo Posicionamento Regional Dec 25" / "PARA_TESTE"
BASE       = PARA_TESTE / "1ª RODADA"

SOURCES = {
    "DEZEMBRO": {
        "file":   BASE / "1. DEZEMBRO" / "DEZEMBRO _ RELATÓRIO CONSOLIDADO _ POSICIONAMENTO ESTADUAL RJ.xlsx",
        "output": PARA_TESTE / "Consolidado Verification - DEZEMBRO - RJ.xlsx",
    },
    "JANEIRO": {
        "file":   BASE / "2. JANEIRO" / "JANEIRO _ RELATÓRIO CONSOLIDADO _ POSICIONAMENTO ESTADUAL RJ.xlsx",
        "output": PARA_TESTE / "Consolidado Verification - JANEIRO - RJ.xlsx",
    },
    "GERAL": {
        "file":   BASE / "3. GERAL" / "GERAL _ RELATÓRIO CONSOLIDADO _ POSICIONAMENTO ESTADUAL RJ.xlsx",
        "output": PARA_TESTE / "Consolidado Verification - GERAL - RJ.xlsx",
    },
}

TEMPLATE_DATA_ROW = 9   # primeira linha de dados no template


# ── Helpers ───────────────────────────────────────────────────────────────────
def _to_int(v) -> int:
    try:
        return int(float(str(v).replace(",", ".")))
    except (ValueError, TypeError):
        return 0


def _to_float(v):
    try:
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _parse_periodo(s) -> tuple[str, str]:
    """'23/12/2025 - 31/12/2025' → ('23/12/2025', '31/12/2025')"""
    s = str(s).strip()
    if not s or s == "nan":
        return "", ""
    parts = s.split(" - ")
    return (parts[0].strip(), parts[1].strip()) if len(parts) == 2 else (s, "")


def _copy_style(src, dst):
    if src.has_style:
        dst.font        = copy(src.font)
        dst.fill        = copy(src.fill)
        dst.border      = copy(src.border)
        dst.alignment   = copy(src.alignment)
        dst.number_format = src.number_format


def _adjust_formula(formula: str, new_row: int) -> str:
    """Substitui referências de linha 9 pelo novo número. Ex: E9 → E10."""
    if not formula or not formula.startswith("="):
        return formula
    return re.sub(r"([A-Z]+)9(?=[^0-9]|$)", lambda m: m.group(1) + str(new_row), formula)


# ── Leitura do RELATÓRIO CONSOLIDADO ─────────────────────────────────────────
def read_source(path: Path) -> tuple[dict, list[dict]]:
    """
    Lê um arquivo RELATÓRIO CONSOLIDADO (formato Nacional/Metrike).

    Retorna:
        metadata: {cliente, agencia, campanha, periodo}
        vehicles: list[dict] com todas as métricas mapeadas para o template
    """
    xl  = pd.ExcelFile(str(path))
    df  = pd.read_excel(xl, sheet_name=xl.sheet_names[0], header=None)

    # ── Metadados (linhas 0–5) ────────────────────────────────────────────────
    meta = {}
    for i in range(6):
        row  = df.iloc[i]
        vals = [str(v).strip() for v in row if str(v).strip() not in ("nan", "")]
        if len(vals) >= 2:
            k = vals[0].rstrip(":").lower()
            if "cliente"  in k: meta["cliente"]  = vals[1]
            elif "agência" in k or "agencia" in k: meta["agencia"]  = vals[1]
            elif "campanha" in k: meta["campanha"] = vals[1]
            elif "período" in k or "periodo" in k: meta["periodo"]  = vals[1]

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    header_row = None
    for i, row in df.iterrows():
        lows = [str(v).strip().lower() for v in row if str(v).strip() not in ("nan", "")]
        if "veículo" in lows or "veiculo" in lows:
            header_row = i
            break

    if header_row is None:
        return meta, []

    df.columns = [
        str(v).strip() if str(v).strip() not in ("nan",) else f"col_{j}"
        for j, v in enumerate(df.iloc[header_row])
    ]
    data = df.iloc[header_row + 1:].reset_index(drop=True)
    data = data[data["Veículo"].notna() & (data["Veículo"].astype(str).str.strip() != "")]

    # ── Veículos ──────────────────────────────────────────────────────────────
    vehicles = []
    for _, row in data.iterrows():
        veiculo = str(row.get("Veículo", "")).strip()
        if not veiculo:
            continue

        tipo   = str(row.get("Tipo de compra", "")).strip().upper()
        is_cpv = tipo == "CPV"

        impressoes = _to_int(row.get("Impressões")) if not is_cpv else 0
        views      = _to_int(row.get("Views"))      if is_cpv     else 0

        data_ini, data_fim = _parse_periodo(row.get("Período", ""))

        # Indevidas — fonte tem agregado; mapeamento:
        #   Impróprio     → col 14 (Acidente) como proxy do agregado
        #   Safeframe     → col 18
        #   Indeterminados → col 21 (Não Classificado)
        improprio      = _to_int(row.get("Impróprio"))
        safeframe      = _to_int(row.get("Safeframe"))
        indeterminados = _to_int(row.get("Indeterminados"))

        devolutiva = str(row.get("Devolutiva BI SECOM", "") or "").strip()
        if devolutiva in ("nan", "-"):
            devolutiva = ""

        vehicles.append({
            "veiculo":          veiculo,
            "tipo_compra":      tipo,
            "contratado":       _to_int(row.get("Contratado")),
            "impressoes":       impressoes,
            "cliques":          _to_int(row.get("Cliques")),
            "views":            views,
            "viewables":        _to_int(row.get("Viewables")),
            "viewability":      _to_float(row.get("Viewability")),
            "entregas_validas": _to_int(row.get("Entregas Válidas")),
            "acidente":         improprio,      # agregado — sem breakdown individual
            "safeframe":        safeframe,
            "nao_classificado": indeterminados,
            "data_ini":         data_ini,
            "data_fim":         data_fim,
            "devolutiva_bi":    devolutiva,
        })

    return meta, vehicles


# ── Escrita do Consolidado Verification ──────────────────────────────────────
def write_consolidado(meta: dict, vehicles: list[dict], output: Path):
    """Copia o template e preenche os dados dos veículos."""
    shutil.copy2(str(TEMPLATE), str(output))
    wb = openpyxl.load_workbook(str(output))
    ws = wb.active

    # ── Metadados (linhas 1–4) ────────────────────────────────────────────────
    ws.cell(row=1, column=2).value = meta.get("cliente",  "SECOM")
    ws.cell(row=2, column=2).value = meta.get("agencia",  "")
    ws.cell(row=3, column=2).value = meta.get("campanha", "")
    ws.cell(row=4, column=2).value = meta.get("periodo",  "")

    # ── Linhas de dados ───────────────────────────────────────────────────────
    for offset, v in enumerate(vehicles):
        r = TEMPLATE_DATA_ROW + offset

        if offset > 0:
            # Copiar formatação e fórmulas da linha 9 para a nova linha
            for col in range(1, 29):
                src = ws.cell(row=TEMPLATE_DATA_ROW, column=col)
                dst = ws.cell(row=r, column=col)
                _copy_style(src, dst)
                if isinstance(src.value, str) and src.value.startswith("="):
                    dst.value = _adjust_formula(src.value, r)

        # Colunas de input (sobrescrevem fórmulas/vazios)
        def w(col, val):
            cell = ws.cell(row=r, column=col)
            cell.value = val if val not in (0, None, "", "nan") or col in (4,) else None
            # Col 4 (Contratado) sempre escreve, mesmo se 0

        ws.cell(row=r, column=1).value  = v["veiculo"]
        ws.cell(row=r, column=2).value  = "RJ"
        ws.cell(row=r, column=3).value  = v["tipo_compra"]
        ws.cell(row=r, column=4).value  = v["contratado"]
        ws.cell(row=r, column=5).value  = v["impressoes"] or None
        # col 6  = fórmula % Entregue
        ws.cell(row=r, column=7).value  = v["cliques"]   or None
        # col 8  = fórmula CTR
        ws.cell(row=r, column=9).value  = v["views"]     or None
        # col 10 = fórmula VTR
        ws.cell(row=r, column=11).value = v["viewables"] or None
        ws.cell(row=r, column=12).value = v["viewability"]
        ws.cell(row=r, column=13).value = v["entregas_validas"] or None
        ws.cell(row=r, column=14).value = v["acidente"]         or None  # Impróprio (agregado)
        ws.cell(row=r, column=15).value = None   # Violência — sem dado
        ws.cell(row=r, column=16).value = None   # Língua Estrangeira — sem dado
        ws.cell(row=r, column=17).value = None   # Pornografia — sem dado
        ws.cell(row=r, column=18).value = v["safeframe"]        or None
        ws.cell(row=r, column=19).value = None   # App Móvel — sem dado
        ws.cell(row=r, column=20).value = None   # Teste de Tag — sem dado
        ws.cell(row=r, column=21).value = v["nao_classificado"] or None
        # col 22 = fórmula % Indevidas
        # col 23 = fórmula Total Indevidas
        # col 24 = fórmula Dif%
        ws.cell(row=r, column=25).value = v["data_ini"] or None
        ws.cell(row=r, column=26).value = v["data_fim"] or None
        ws.cell(row=r, column=27).value = v["devolutiva_bi"] or None

    wb.save(str(output))
    wb.close()
    print(f"  ✓  {output.name}  ({len(vehicles)} veículo{'s' if len(vehicles) != 1 else ''})")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not TEMPLATE.exists():
        print(f"ERRO: template não encontrado em {TEMPLATE}")
        return

    print("Gerando consolidados...\n")
    for nome, cfg in SOURCES.items():
        src = cfg["file"]
        if not src.exists():
            print(f"  ✗  {nome}: arquivo não encontrado — {src.name}")
            continue
        meta, vehicles = read_source(src)
        if not vehicles:
            print(f"  ✗  {nome}: nenhum veículo lido")
            continue
        write_consolidado(meta, vehicles, cfg["output"])

    print("\nConcluído.")


if __name__ == "__main__":
    main()
