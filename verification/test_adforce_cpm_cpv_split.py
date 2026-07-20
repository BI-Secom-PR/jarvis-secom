#!/usr/bin/env python3
"""Self-check: ADFORCE consolidado com 2 linhas do mesmo veículo, uma CPM e
outra CPV (caso real: F5 ADS em CONSOLIDADO_ABRIL, cada objetivo num arquivo
comprovante E num arquivo verification separado). Sem isso, _merge_by_veiculo()
somava os dois objetivos num total único tanto do lado do comprovante (entregue/
views/cliques) quanto do verification (indevidas e o total usado no resumo DIF),
e ambas as linhas do consolidado eram comparadas contra o mesmo total combinado —
gerando DIVERGENCIA falsa e um "DIF impressoes/views" idêntico nas duas linhas.

Constrói fixtures mínimas em memória (não depende dos arquivos reais do
usuário, que ficam fora do repo) e roda engine.py de ponta a ponta.

Uso:  python3 verification/test_adforce_cpm_cpv_split.py
"""
import json
import os
import subprocess
import sys
import tempfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ENGINE = os.path.join(HERE, "..", "app", "verification", "engine.py")

VEICULO = "Rádio Teste"


def _build_comp_cpm(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Veículo", "Impressões", "Cliques", "Tipo de Compra"])
    ws.append([VEICULO, 100000, 500, "CPM"])
    wb.save(path)


def _build_comp_cpv(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Veículo", "Impressões", "Cliques", "0%", "50%", "100%", "Tipo de Compra"])
    ws.append([VEICULO, 20000, 100, 18000, 15000, 12000, "CPV"])
    wb.save(path)


def _build_verif_cpm(path: str) -> None:
    """Flat ADFORCE, arquivo puro-CPM (caso real: 1 verif file por objetivo)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["vehicle", "categories", "url", "cpm"])
    ws.append([VEICULO, "Safeframe", "http://sitea.com/1", 300])
    ws.append([VEICULO, "Safeframe", "http://sitea.com/2", 200])
    wb.save(path)


def _build_verif_cpv(path: str) -> None:
    """Flat ADFORCE, arquivo puro-CPV."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["vehicle", "categories", "url", "cpv"])
    ws.append([VEICULO, "Safeframe", "http://siteb.com/1", 400])
    wb.save(path)


def _build_consolidado(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verification"
    # Linhas 1-7: metadados (ignorados pelo engine, só HEADER_ROW=8/DATA_START_ROW=9 importam)
    for _ in range(7):
        ws.append([])
    header = [""] * 28
    header[0] = "Veículo"          # col 1 = COL_VEICULO
    header[2] = "Objetivo de Mídia"  # col 3 = COL_TIPO (posicional, texto é só cosmético)
    header[3] = "Contratado"       # col 4 = COL_CONTRATADO
    header[4] = "Impressões"       # col 5 = COL_IMPRESSOES
    header[6] = "Cliques"          # col 7 = COL_CLIQUES
    header[8] = "Views"            # col 9 = COL_VIEWS
    header[18] = "Safeframe"       # col 19 = COL_INDEVIDAS["safeframe"]
    ws.append(header)

    row_cpm = [""] * 28
    row_cpm[0], row_cpm[2], row_cpm[3] = VEICULO, "CPM", 100000
    row_cpm[4], row_cpm[6] = 100000, 500
    row_cpm[18] = 500
    ws.append(row_cpm)

    row_cpv = [""] * 28
    row_cpv[0], row_cpv[2], row_cpv[3] = VEICULO, "CPV", 20000
    # Views (col 9) espera views iniciadas (0%), não completadas (100%) — 18000 é
    # o valor de "0%" no comprovante CPV abaixo, não o de "100%".
    row_cpv[4], row_cpv[6], row_cpv[8] = 20000, 100, 18000
    row_cpv[18] = 400
    ws.append(row_cpv)

    wb.save(path)


def main() -> None:
    failures: list[str] = []
    with tempfile.TemporaryDirectory() as outdir:
        comp_cpm = os.path.join(outdir, "comp_cpm.xlsx")
        comp_cpv = os.path.join(outdir, "comp_cpv.xlsx")
        verif_cpm = os.path.join(outdir, "verif_cpm.xlsx")
        verif_cpv = os.path.join(outdir, "verif_cpv.xlsx")
        consolidado = os.path.join(outdir, "consolidado.xlsx")
        _build_comp_cpm(comp_cpm)
        _build_comp_cpv(comp_cpv)
        _build_verif_cpm(verif_cpm)
        _build_verif_cpv(verif_cpv)
        _build_consolidado(consolidado)

        cmd = [
            sys.executable, ENGINE, consolidado,
            "--adserver", "adforce",
            "--comp", comp_cpm, comp_cpv,
            "--verif", verif_cpm, verif_cpv,
            "--url-pct", "100",
            "--output", os.path.join(outdir, "out.xlsx"),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True,
                              cwd=os.path.join(HERE, "..", "app", "verification"))
        assert proc.returncode == 0, f"engine exit {proc.returncode}\n{proc.stderr[-2000:]}"
        r = json.loads(proc.stdout)

        if r["parse_errors"]:
            failures.append(f"parse_errors: {r['parse_errors']}")

        veiculos = r["veiculos"]
        if len(veiculos) != 2:
            failures.append(f"esperava 2 linhas (CPM+CPV), veio {len(veiculos)}: {veiculos}")
        else:
            cpm_row = next((v for v in veiculos if "OK entregue: 100.000" in v["devolutiva"]), None)
            cpv_row = next((v for v in veiculos if "OK entregue: 20.000" in v["devolutiva"]), None)

            if cpm_row is None:
                failures.append(f"linha CPM não bateu entregue=100.000 (blend teria dado 120.000): {veiculos}")
            else:
                if cpm_row["status"] != "OK":
                    failures.append(f"linha CPM status={cpm_row['status']} (esperava OK): {cpm_row['devolutiva']}")
                if "OK safeframe: 500" not in cpm_row["devolutiva"]:
                    failures.append(f"linha CPM safeframe != 500 (indevidas_cpm não isolado): {cpm_row['devolutiva']}")
                # DIF: verif do arquivo CPM (500) isolado, não o blend com o CPV (900)
                if "verif 500" not in cpm_row["devolutiva"]:
                    failures.append(f"linha CPM DIF usou verif blendado (esperava 'verif 500'): {cpm_row['devolutiva']}")

            if cpv_row is None:
                failures.append(f"linha CPV não bateu entregue=20.000 (blend teria dado 120.000): {veiculos}")
            else:
                if cpv_row["status"] != "OK":
                    failures.append(f"linha CPV status={cpv_row['status']} (esperava OK): {cpv_row['devolutiva']}")
                if "OK views: 18.000" not in cpv_row["devolutiva"]:
                    failures.append(f"linha CPV views != 18.000: {cpv_row['devolutiva']}")
                if "OK safeframe: 400" not in cpv_row["devolutiva"]:
                    failures.append(f"linha CPV safeframe != 400 (indevidas_cpv não isolado): {cpv_row['devolutiva']}")
                # DIF: verif do arquivo CPV (400) isolado, não o blend com o CPM (900)
                if "verif 400" not in cpv_row["devolutiva"]:
                    failures.append(f"linha CPV DIF usou verif blendado (esperava 'verif 400'): {cpv_row['devolutiva']}")

        # tipo_by_verif_norm: com 2 tipos no mesmo veículo, filtro de objetivo deve
        # ser desativado — URLs de origem CPM (sitea.com) E CPV (siteb.com) devem
        # aparecer no pool, sem a última linha processada sobrescrever a outra.
        urls = {item["url"] for item in r["url_sample"]}
        if "http://siteb.com/1" not in urls:
            failures.append(f"URL CPV (siteb.com) ausente do url_sample: {urls}")
        if not (urls & {"http://sitea.com/1", "http://sitea.com/2"}):
            failures.append(f"URLs CPM (sitea.com) ausentes do url_sample — last-write-wins voltou: {urls}")

    if failures:
        print(f"\n{len(failures)} FALHA(S):")
        for f in failures:
            print(f"  - {f}")
        sys.exit(1)
    print("OK — CPM e CPV do mesmo veículo comparados contra seus próprios subtotais.")


if __name__ == "__main__":
    main()
