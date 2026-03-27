"""
Parser 00px — comprovante de entrega + verification de URLs.

Comprovante:
  Workbook com sheets "CPM - Contabilizações", "CPC - Contabilizações",
  "CPV - Contabilizações". Cada sheet tem metadados nas primeiras linhas
  seguidos de uma tabela com cabeçalho detectado automaticamente.
  Viewability: coluna "VA (IAB)" (pode ser %, ex.: "72%", ou decimal 0.72).

Verification:
  Sheet "Verification" (ou wb.active). Colunas: Veículo, Categoria, URL.
  Indevidas mapeadas via category_map.
"""

import json
import random
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from category_map import INDEVIDAS_ZERO, normaliza_categoria
from parser_utils import col_index, parse_date, to_float, to_int, cli_date

# ── Helpers ─────────────────────────────────────────────────────────────────────

def _normalize_va(v) -> float | None:
    """Converte VA decimal (0.72) → percentagem (72.0); mantém se já > 1."""
    f = to_float(v)
    if f is None:
        return None
    return round(f * 100, 2) if f <= 1.0 else round(f, 2)


def _find_header(ws, veiculo_required: bool = True):
    """
    Varre até 25 linhas buscando cabeçalho.
    Se veiculo_required=True, exige col com 'Veículo'/'Vehicle'.
    Sempre exige col com 'Impressões'/'Impressions'/'Entregues'/'Views'.
    Retorna (row_idx_1based, header_list) ou (None, []).
    """
    veiculo_variants  = {"veículo", "veiculo", "vehicle"}
    entregue_variants = {"impressões", "impressoes", "impressions",
                         "entregues", "entregue", "views"}
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        has_entregue = bool(lows & entregue_variants)
        has_veiculo  = bool(lows & veiculo_variants)
        if has_entregue and (has_veiculo or not veiculo_required):
            return i, vals
    return None, []


def _get_comp_sheets(wb):
    """Retorna sheets de Contabilizações (CPM/CPC/CPV). Fallback: primeira sheet."""
    targeted = [
        ws for ws in wb.worksheets
        if "contabiliza" in ws.title.lower()
    ]
    return targeted or [wb.worksheets[0]]


SKIP_VEHICLE_NAMES = {"---", "grand total:", "grand total", "total"}


# ── parse_comprovante ────────────────────────────────────────────────────────────

def parse_comprovante(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia comprovante 00px (multi-sheet CPM/CPC/CPV)."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = _get_comp_sheets(wb)

    entregue:    dict[str, int]   = defaultdict(int)
    cliques:     dict[str, int]   = defaultdict(int)
    viewables:   dict[str, int]   = defaultdict(int)
    contratado:  dict[str, int]   = {}
    viewability: dict[str, float] = {}
    va_wsum:     dict[str, float] = defaultdict(float)
    va_weight:   dict[str, int]   = defaultdict(int)
    last_vehicle: str | None      = None
    found_header  = False

    for ws in sheets:
        header_row_idx, header = _find_header(ws)
        if header_row_idx is None:
            continue
        found_header = True

        i_veiculo    = col_index(header, "Veículo", "Veiculo", "Vehicle")
        i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions",
                                  "Entregues", "Entregue", "Views")
        i_cliques    = col_index(header, "Cliques", "Clicks")
        i_viewable   = col_index(header, "Viewable", "Viewables")
        i_viewability= col_index(header, "VA%", "VA (IAB)", "Viewability", "VA %",
                                  "View%", "Viewability (IAB)")
        i_contratado = col_index(header, "Contratado", "Contracted")
        i_data       = col_index(header, "Data", "Date")

        if i_veiculo is None or i_impressoes is None:
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(v is None for v in row):
                continue

            veiculo_raw = row[i_veiculo] if i_veiculo < len(row) else None
            if veiculo_raw is None:
                continue
            vname = str(veiculo_raw).strip()
            if not vname:
                continue
            if vname.lower() in SKIP_VEHICLE_NAMES:
                continue
            if vname.endswith(" Total") or vname.endswith(" total"):
                continue

            # Linha #TOTAL → contratado e viewability do veículo anterior
            if "#TOTAL" in vname.upper():
                if last_vehicle is not None:
                    if i_contratado is not None and i_contratado < len(row):
                        c = to_int(row[i_contratado])
                        if c > 0:
                            contratado[last_vehicle] = c
                    if i_viewability is not None and i_viewability < len(row):
                        va = _normalize_va(row[i_viewability])
                        if va is not None:
                            viewability[last_vehicle] = va
                continue

            # Filtro de data
            if i_data is not None and i_data < len(row):
                raw_date = row[i_data]
                if raw_date is not None:
                    d = parse_date(raw_date)
                    if d is None:
                        # Unparseable value in date column (e.g. "#TOTAL POR VEÍCULO",
                        # "#TOTAL POR CANAL") → subtotal row, skip to avoid double-count
                        continue
                    if data_ini and d < data_ini:
                        continue
                    if data_fim and d > data_fim:
                        continue

            imp = to_int(row[i_impressoes] if i_impressoes < len(row) else None)
            if imp == 0:
                continue

            entregue[vname] += imp
            if i_cliques is not None and i_cliques < len(row):
                cliques[vname] += to_int(row[i_cliques])
            if i_viewable is not None and i_viewable < len(row):
                viewables[vname] += to_int(row[i_viewable])
            if i_viewability is not None and i_viewability < len(row):
                va = _normalize_va(row[i_viewability])
                if va is not None and imp > 0:
                    va_wsum[vname]   += va * imp
                    va_weight[vname] += imp

            last_vehicle = vname

    wb.close()

    if not found_header:
        raise ValueError(
            f"Cabeçalho com 'Veículo' e 'Impressões' não encontrado em nenhuma sheet: {path.name}"
        )
    if not entregue:
        return []

    # Preencher viewability pela média ponderada onde não há linha #TOTAL
    for vname in entregue:
        if vname not in viewability and va_weight[vname] > 0:
            viewability[vname] = round(va_wsum[vname] / va_weight[vname], 2)

    return [
        {
            "veiculo":           vname,
            "tipo_compra":       None,
            "contratado":        contratado.get(vname),
            "entregue":          imp,
            "cliques":           cliques[vname] or None,
            "viewables":         viewables[vname] or None,
            "viewability":       viewability.get(vname),
            "indevidas":         {},
            "url_sample":        [],
            "formato_detectado": "00px_comprovante",
        }
        for vname, imp in entregue.items()
    ]


# ── parse_verif ──────────────────────────────────────────────────────────────────

def parse_verif(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia arquivo de verification 00px (URL + categoria por linha)."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Detectar header: exige 'categoria' + alguma col com 'url'
    header_row_idx, header = None, []
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if "categoria" in lows and lows & {"veículo", "veiculo", "vehicle", "veículos", "veiculos"}:
            header_row_idx, header = i, vals
            break

    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Header com 'Categoria' e 'URL' não encontrado nas primeiras 25 linhas: {path.name}"
        )

    i_veiculo    = col_index(header, "Veículo", "Veiculo", "Vehicle",
                              "Veículos", "Veiculos")
    i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions",
                              "Impressões Totais", "Impressoes Totais")
    # CPV rows store delivery count in "Views" — CPM rows leave it blank (and vice-versa)
    i_views      = col_index(header, "Views", "Visualizações", "Visualizacoes")
    i_categoria  = col_index(header, "Categoria", "Category")
    i_url        = col_index(header, "Url", "URL", "url", "URL Veiculada",
                              "Url Veiculada")
    i_data       = col_index(header, "Data", "Date")

    if i_veiculo is None or (i_impressoes is None and i_views is None) or i_categoria is None:
        wb.close()
        raise ValueError(
            f"Colunas obrigatórias ausentes (Veículo/Impressões ou Views/Categoria): {path.name}"
        )

    veiculos_indevidas: dict[str, dict] = defaultdict(lambda: dict(INDEVIDAS_ZERO))
    veiculos_entregue:  dict[str, int]  = defaultdict(int)

    MAX_POOL = 500
    url_pool: list[dict] = []
    pool_count = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue
        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        veiculo   = str(row[i_veiculo]).strip()   if i_veiculo   < len(row) and row[i_veiculo]   else None
        categoria = str(row[i_categoria]).strip()  if i_categoria < len(row) and row[i_categoria] else None
        url       = str(row[i_url]).strip()        if i_url is not None and i_url < len(row) and row[i_url] else None
        # CPM rows fill Impressões; CPV rows fill Views — use whichever is non-zero
        impressoes = (
            to_int(row[i_impressoes] if i_impressoes is not None and i_impressoes < len(row) else None)
            or to_int(row[i_views]      if i_views      is not None and i_views      < len(row) else None)
        )

        if not veiculo or not categoria:
            continue

        cat_key = normaliza_categoria(categoria)
        if cat_key:
            veiculos_indevidas[veiculo][cat_key] = (
                veiculos_indevidas[veiculo].get(cat_key, 0) + impressoes
            )

        veiculos_entregue[veiculo] += impressoes

        if url and cat_key:
            pool_count += 1
            entry = {"url": url, "categoria": categoria, "veiculo": veiculo}
            if len(url_pool) < MAX_POOL:
                url_pool.append(entry)
            else:
                idx = random.randint(0, pool_count - 1)
                if idx < MAX_POOL:
                    url_pool[idx] = entry

    wb.close()

    results: list[dict] = []
    for veiculo in veiculos_entregue:
        results.append({
            "veiculo":           veiculo,
            "tipo_compra":       None,
            "contratado":        None,
            "entregue":          veiculos_entregue[veiculo],
            "cliques":           None,
            "viewables":         None,
            "viewability":       None,
            "indevidas":         dict(veiculos_indevidas[veiculo]),
            "url_sample":        url_pool if not results else [],
            "formato_detectado": "00px_verif",
        })

    return results


# ── CLI ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parser 00px")
    ap.add_argument("modo", choices=["comp", "verif"],
                    help="'comp' = comprovante, 'verif' = verification URL")
    ap.add_argument("arquivo", help="Arquivo .xlsx")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    fn = parse_comprovante if args.modo == "comp" else parse_verif
    try:
        res = fn(args.arquivo, data_ini=cli_date(args.ini), data_fim=cli_date(args.fim))
        for r in res:
            disp = {k: v for k, v in r.items() if k != "url_sample"}
            disp["url_sample_count"] = len(r.get("url_sample", []))
            print(json.dumps(disp, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, ensure_ascii=False))
        sys.exit(1)
