"""
Parser AHEAD — comprovante de entrega + verification de URLs.

Comprovante:
  Mesmo formato DCM/CM360 do ADMOTION. Sheet "Data" com cabeçalho
  em torno da linha 15 (metadados pesados antes). Usa "Site (CM360)"
  como coluna de veículo.

Verification:
  Sheet "VERIFICATION". Colunas: Veículos (plural), Data, URL Veiculada,
  Impressões Totais, Categoria.
  Nota: categorias AHEAD são tipicamente "valid-request" → sem indevidas.
"""

import json
import random
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from category_map import INDEVIDAS_ZERO, normaliza_categoria
from parser_utils import col_index, parse_date, to_float, to_int, cli_date, vehicle_from_filename


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _normalize_va(v) -> float | None:
    f = to_float(v)
    if f is None:
        return None
    return round(f * 100, 2) if f <= 1.0 else round(f, 2)


SKIP_VEHICLE_NAMES = {"---", "grand total:", "grand total", "total"}


def _find_comp_header(ws):
    """Varre até 25 linhas: 'Site (CM360)' OU 'Veículo' E 'Impressions'."""
    veiculo_variants  = {"site (cm360)", "veículo", "veiculo", "vehicle"}
    entregue_variants = {"impressions", "impressões", "impressoes",
                         "entregues", "entregue"}
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if lows & veiculo_variants and lows & entregue_variants:
            return i, vals
    return None, []


def _find_verif_header(ws):
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if "categoria" in lows and lows & {"veículo", "veiculo", "vehicle", "veículos", "veiculos"}:
            return i, vals
    return None, []


# ── parse_comprovante ────────────────────────────────────────────────────────────

def parse_comprovante(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia comprovante AHEAD (formato CM360)."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    data_ws = next(
        (ws for ws in wb.worksheets if ws.title.strip().lower() == "data"),
        wb.worksheets[0]
    )

    header_row_idx, header = _find_comp_header(data_ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Cabeçalho com 'Site (CM360)' e 'Impressions' não encontrado: {path.name}"
        )

    i_veiculo    = col_index(header, "Site (CM360)", "Veículo", "Veiculo", "Vehicle")
    i_impressoes = col_index(header, "Impressions", "Impressões", "Impressoes",
                              "Entregues", "Entregue")
    i_cliques    = col_index(header, "Clicks", "Cliques")
    i_viewable   = col_index(header, "Active View: Viewable Impressions",
                              "Viewable", "Viewables")
    i_viewability= col_index(header, "Active View: % Viewable Impressions",
                              "VA%", "Viewability", "VA (IAB)")
    i_contratado = col_index(header, "Flight Booked Units", "Contratado", "Contracted")
    i_data       = col_index(header, "Date", "Data")

    if i_impressoes is None:
        wb.close()
        raise ValueError(f"Coluna obrigatória ausente (Impressions): {path.name}")
    fallback_vehicle = vehicle_from_filename(filepath)

    entregue:   dict[str, int]   = defaultdict(int)
    cliques:    dict[str, int]   = defaultdict(int)
    viewables:  dict[str, int]   = defaultdict(int)
    contratado: dict[str, int]   = {}
    viewability:dict[str, float] = {}
    va_wsum:    dict[str, float] = defaultdict(float)
    va_weight:  dict[str, int]   = defaultdict(int)

    for row in data_ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        veiculo_raw = row[i_veiculo] if i_veiculo is not None and i_veiculo < len(row) else None
        vname = str(veiculo_raw).strip() if veiculo_raw is not None else ""
        if not vname or vname.lower() in SKIP_VEHICLE_NAMES:
            vname = fallback_vehicle
        if vname.endswith(" Total") or vname.endswith(" total"):
            continue

        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        imp = to_int(row[i_impressoes] if i_impressoes < len(row) else None)
        if imp == 0:
            continue

        entregue[vname] += imp
        if i_cliques is not None and i_cliques < len(row):
            cliques[vname] += to_int(row[i_cliques])
        if i_viewable is not None and i_viewable < len(row):
            viewables[vname] += to_int(row[i_viewable])
        if i_viewability is not None and i_viewability < len(row):
            va = _normalize_va(row[i_viewability])
            if va is not None and imp > 0:
                va_wsum[vname]   += va * imp
                va_weight[vname] += imp
        if i_contratado is not None and i_contratado < len(row):
            c = to_int(row[i_contratado])
            if c > 0 and vname not in contratado:
                contratado[vname] = c

    wb.close()

    if not entregue:
        return []

    for vname in entregue:
        if vname not in viewability and va_weight[vname] > 0:
            viewability[vname] = round(va_wsum[vname] / va_weight[vname], 2)

    return [
        {
            "veiculo":           vname,
            "tipo_compra":       None,
            "contratado":        contratado.get(vname),
            "entregue":          imp,
            "cliques":           cliques[vname] or None,
            "viewables":         viewables[vname] or None,
            "viewability":       viewability.get(vname),
            "indevidas":         {},
            "url_sample":        [],
            "formato_detectado": "ahead_comprovante",
        }
        for vname, imp in entregue.items()
    ]


# ── parse_verif ──────────────────────────────────────────────────────────────────

def parse_verif(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia verification AHEAD (URL Veiculada, Veículos, Impressões Totais)."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    header_row_idx, header = _find_verif_header(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Header com 'Categoria' e 'URL' não encontrado nas primeiras 25 linhas: {path.name}"
        )

    i_veiculo    = col_index(header, "Veículos", "Veiculos", "Veículo", "Veiculo",
                              "Vehicle")
    i_impressoes = col_index(header, "Impressões Totais", "Impressoes Totais",
                              "Impressões", "Impressoes", "Impressions")
    i_categoria  = col_index(header, "Categoria", "Category")
    i_url        = col_index(header, "URL Veiculada", "Url Veiculada",
                              "Url", "URL", "url")
    i_data       = col_index(header, "Data", "Date")

    if i_veiculo is None or i_impressoes is None or i_categoria is None:
        wb.close()
        raise ValueError(
            f"Colunas obrigatórias ausentes (Veículos/Impressões/Categoria): {path.name}"
        )

    veiculos_indevidas: dict[str, dict] = defaultdict(lambda: dict(INDEVIDAS_ZERO))
    veiculos_entregue:  dict[str, int]  = defaultdict(int)
    MAX_POOL = 500
    url_pool: list[dict] = []
    pool_count = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue
        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        veiculo   = str(row[i_veiculo]).strip()   if i_veiculo   < len(row) and row[i_veiculo]   else None
        categoria = str(row[i_categoria]).strip()  if i_categoria < len(row) and row[i_categoria] else None
        url       = str(row[i_url]).strip()        if i_url is not None and i_url < len(row) and row[i_url] else None
        impressoes = to_int(row[i_impressoes] if i_impressoes < len(row) else None)

        if not veiculo or not categoria:
            continue

        cat_key = normaliza_categoria(categoria)
        if cat_key:
            veiculos_indevidas[veiculo][cat_key] = (
                veiculos_indevidas[veiculo].get(cat_key, 0) + impressoes
            )

        veiculos_entregue[veiculo] += impressoes

        if url and cat_key:
            pool_count += 1
            entry = {"url": url, "categoria": categoria, "veiculo": veiculo}
            if len(url_pool) < MAX_POOL:
                url_pool.append(entry)
            else:
                idx = random.randint(0, pool_count - 1)
                if idx < MAX_POOL:
                    url_pool[idx] = entry

    wb.close()

    results: list[dict] = []
    for veiculo in veiculos_entregue:
        results.append({
            "veiculo":           veiculo,
            "tipo_compra":       None,
            "contratado":        None,
            "entregue":          veiculos_entregue[veiculo],
            "cliques":           None,
            "viewables":         None,
            "viewability":       None,
            "indevidas":         dict(veiculos_indevidas[veiculo]),
            "url_sample":        url_pool if not results else [],
            "formato_detectado": "ahead_verif",
        })

    return results


# ── CLI ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parser AHEAD")
    ap.add_argument("modo", choices=["comp", "verif"])
    ap.add_argument("arquivo")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    fn = parse_comprovante if args.modo == "comp" else parse_verif
    try:
        res = fn(args.arquivo, data_ini=cli_date(args.ini), data_fim=cli_date(args.fim))
        for r in res:
            disp = {k: v for k, v in r.items() if k != "url_sample"}
            disp["url_sample_count"] = len(r.get("url_sample", []))
            print(json.dumps(disp, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, ensure_ascii=False))
        sys.exit(1)
