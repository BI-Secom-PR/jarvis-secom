"""
Parser genérico para arquivos de comprovante de entrega.

Detectado por: nome de arquivo contém 'comprovante' (qualquer case).

Estrutura esperada (primeira aba, independente do nome):
  - Linhas 1–N: metadados livres
  - Linha M:    cabeçalho com ao menos 'Veículo' e 'Impressões' (ou variantes)
  - Linha M+1+: dados de entrega por placement/dia, intercalados com
                linhas de total "#TOTAL POR VEÍCULO" (quando formato Nacional)

Saída (list[dict] — um dict por veículo encontrado):
  {
    veiculo:           str,
    tipo_compra:       None,
    contratado:        int | None,
    entregue:          int,          # soma de Impressões das linhas de dados
    cliques:           int | None,
    viewables:         int | None,
    viewability:       float | None, # VA% da linha #TOTAL correspondente
    indevidas:         {},            # sempre vazio — comprovante não tem categorias
    url_sample:        [],
    formato_detectado: "comprovante",
  }
"""

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from parser_utils import col_index, parse_date, to_float, to_int, cli_date


def _find_header(ws) -> tuple[int | None, list[str]]:
    """
    Varre as primeiras 25 linhas procurando cabeçalho com 'veículo' E
    ('impressões' ou 'impressoes' ou 'entregues').
    Retorna (row_idx_1based, header_values) ou (None, []).
    """
    veiculo_variants  = {"veículo", "veiculo", "vehicle"}
    entregue_variants = {"impressões", "impressoes", "impressions", "entregues", "entregue"}

    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if lows & veiculo_variants and lows & entregue_variants:
            return i, vals
    return None, []


def parse(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """
    Parseia um arquivo de comprovante de entrega (primeira aba).

    Retorna list[dict] — um dict por veículo.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]  # sempre primeira aba

    header_row_idx, header = _find_header(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Cabeçalho com 'Veículo' e 'Impressões' não encontrado nas primeiras 25 linhas: {path.name}"
        )

    # Índices de colunas (0-based)
    i_veiculo    = col_index(header, "Veículo", "Veiculo", "Vehicle")
    i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions", "Entregues", "Entregue")
    i_cliques    = col_index(header, "Cliques", "Clicks")
    i_viewable   = col_index(header, "Viewable", "Viewables")
    i_viewability= col_index(header, "VA%", "Viewability", "VA %", "View%")
    i_contratado = col_index(header, "Contratado", "Contracted")
    i_data       = col_index(header, "Data", "Date")

    if i_veiculo is None or i_impressoes is None:
        wb.close()
        raise ValueError(
            f"Colunas obrigatórias ausentes (Veículo/Impressões): {path.name}"
        )

    # Acumuladores por veículo
    entregue:    dict[str, int]          = defaultdict(int)
    cliques:     dict[str, int]          = defaultdict(int)
    viewables:   dict[str, int]          = defaultdict(int)
    contratado:  dict[str, int]          = {}
    viewability: dict[str, float]        = {}
    last_vehicle: str | None             = None  # veículo real antes da linha #TOTAL

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        veiculo_raw = row[i_veiculo] if i_veiculo < len(row) else None
        if veiculo_raw is None:
            continue
        vname = str(veiculo_raw).strip()
        if not vname:
            continue

        # Linha de total (#TOTAL POR VEÍCULO): extrair Contratado e VA%
        # associados ao último veículo real visto antes desta linha
        if "#TOTAL" in vname.upper():
            if last_vehicle is not None:
                if i_contratado is not None and i_contratado < len(row):
                    cont_val = to_int(row[i_contratado])
                    if cont_val > 0:
                        contratado[last_vehicle] = cont_val
                if i_viewability is not None and i_viewability < len(row):
                    vab = to_float(row[i_viewability])
                    if vab is not None:
                        viewability[last_vehicle] = vab
            continue

        # Filtro de data
        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        imp = to_int(row[i_impressoes] if i_impressoes < len(row) else None)
        entregue[vname]  += imp
        if i_cliques is not None and i_cliques < len(row):
            cliques[vname]  += to_int(row[i_cliques])
        if i_viewable is not None and i_viewable < len(row):
            viewables[vname] += to_int(row[i_viewable])

        last_vehicle = vname

    wb.close()

    if not entregue:
        return []

    return [
        {
            "veiculo":           vname,
            "tipo_compra":       None,
            "contratado":        contratado.get(vname),
            "entregue":          imp,
            "cliques":           cliques[vname] or None,
            "viewables":         viewables[vname] or None,
            "viewability":       viewability.get(vname),
            "indevidas":         {},
            "url_sample":        [],
            "formato_detectado": "comprovante",
        }
        for vname, imp in entregue.items()
    ]


# ── CLI para testes ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parser genérico de comprovante de entrega")
    ap.add_argument("arquivo", help="Arquivo comprovante_*.xlsx")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    try:
        resultados = parse(
            args.arquivo,
            data_ini=cli_date(args.ini),
            data_fim=cli_date(args.fim),
        )
        print(json.dumps(resultados, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, indent=2, ensure_ascii=False))
        sys.exit(1)
