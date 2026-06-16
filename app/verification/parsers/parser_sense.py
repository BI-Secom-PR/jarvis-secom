"""
Parser SENSE — comprovante de entrega + verification de URLs.

Um arquivo por veículo nos dois formatos.

Comprovante (relatorio_veiculo_*.xlsx):
  Sheet "<TIPO> - Contabilizações" (ex.: "CPM - Contabilizações").
  Metadados nas linhas 2-6, col B: Cliente, Agência, Veículo, Campanha, Período.
  Tabela com cabeçalho ~linha 10: Data, Segmentação, Formato, Dimensão,
  ID Placement, Região, ID Linha Criativa, Linha Criativa, Qtd Contratada,
  Impressões, Bloqueados, Válidas, Usuários Únicos, Viewables, Cliques,
  Viewability. Primeira linha de dados é um "Total" (ignorada — somamos as
  linhas diárias para respeitar o filtro de datas).

Verification (relatorio_verification_*.xlsx):
  Sheet "RELATÓRIO SIMPLIFICADO", mesmo bloco de metadados, cabeçalho ~linha 9:
  Veículo, URL/APP, TAG, Categoria, Impressões válidas, Impressões indevidas,
  Total de impressões. A ordem das colunas varia entre arquivos (PORTAL FORUM
  tem Categoria antes de URL/APP) — posições sempre detectadas pelo cabeçalho.
  Cabeçalhos de impressões embutem totais após '\\n' — só a 1ª linha conta.
"""

import json
import random
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from parser_utils import col_index, parse_date, to_int, cli_date, vehicle_from_filename


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _clean_header(row) -> list[str]:
    """Normaliza células de cabeçalho: só a 1ª linha do texto (totais vêm após '\\n')."""
    return [str(v).split("\n")[0].strip() if v is not None else "" for v in row]


def _veiculo_meta(ws) -> str | None:
    """Lê 'Veículo: <nome>' do bloco de metadados (col B, linhas 1-8)."""
    for row in ws.iter_rows(max_row=8, values_only=True):
        for v in row:
            s = str(v).strip() if v else ""
            if s.lower().startswith(("veículo:", "veiculo:")):
                nome = s.split(":", 1)[1].strip()
                if nome:
                    return nome
    return None


def _find_header(ws, *required: str):
    """Localiza a linha de cabeçalho que contém todos os nomes em `required`."""
    req = {r.lower() for r in required}
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = _clean_header(row)
        lows = {v.lower() for v in vals}
        if req <= lows:
            return i, vals
    return None, []


# ── parse_comprovante ────────────────────────────────────────────────────────────

def parse_comprovante(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia comprovante SENSE somando as linhas diárias de cada sheet
    '<TIPO> - Contabilizações' (a linha 'Total' é ignorada)."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    sheets = [ws for ws in wb.worksheets if "contabiliza" in ws.title.lower()]
    if not sheets:
        sheets = [wb.worksheets[0]]

    veiculo_nome = _veiculo_meta(sheets[0]) or vehicle_from_filename(filepath)

    tipo_compra = None
    contratado = 0
    entregue = 0
    cliques = 0
    viewables = 0

    for ws in sheets:
        if "-" in ws.title and tipo_compra is None:
            tipo_compra = ws.title.split("-")[0].strip() or None

        header_row_idx, header = _find_header(ws, "Data", "Impressões")
        if header_row_idx is None:
            wb.close()
            raise ValueError(
                f"Cabeçalho com 'Data' e 'Impressões' não encontrado: {path.name} ({ws.title})"
            )

        i_data       = col_index(header, "Data")
        i_contratado = col_index(header, "Qtd Contratada", "Contratado")
        i_validas    = col_index(header, "Válidas", "Validas")
        i_impressoes = col_index(header, "Impressões", "Impressoes")
        i_viewables  = col_index(header, "Viewables", "Viewable")
        i_cliques    = col_index(header, "Cliques", "Clicks")

        i_entregue = i_validas if i_validas is not None else i_impressoes
        if i_entregue is None:
            wb.close()
            raise ValueError(f"Coluna obrigatória ausente (Válidas/Impressões): {path.name}")

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            row = list(row)
            if all(v is None for v in row):
                continue
            data_raw = row[i_data] if i_data is not None and i_data < len(row) else None
            data_str = str(data_raw).strip() if data_raw is not None else ""
            if data_str.lower() == "total":
                continue
            d = parse_date(data_raw)
            if d is None:
                continue
            if data_ini and d < data_ini:
                continue
            if data_fim and d > data_fim:
                continue

            if i_contratado is not None and i_contratado < len(row) and row[i_contratado] is not None:
                contratado += to_int(row[i_contratado])
            entregue += to_int(row[i_entregue] if i_entregue < len(row) else None)
            if i_viewables is not None and i_viewables < len(row):
                viewables += to_int(row[i_viewables])
            if i_cliques is not None and i_cliques < len(row):
                cliques += to_int(row[i_cliques])

    wb.close()

    viewability = (
        round(viewables / entregue * 100, 2)
        if viewables > 0 and entregue > 0
        else None
    )

    return [{
        "veiculo":           veiculo_nome,
        "tipo_compra":       tipo_compra,
        "contratado":        contratado or None,
        "entregue":          entregue,
        "cliques":           cliques or None,
        "viewables":         viewables or None,
        "viewability":       viewability,
        "indevidas":         {},
        "url_sample":        [],
        "formato_detectado": "sense_comprovante",
    }]


# ── parse_verif ──────────────────────────────────────────────────────────────────

def parse_verif(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
    praca: str | None = None,
) -> list[dict]:
    """Parseia verification SENSE (Veículo, URL/APP, Categoria, Total de impressões).

    Os arquivos atuais não têm colunas de Data nem Estado; os filtros só são
    aplicados se as colunas existirem (mesmo padrão do METRIKE).
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row_idx, header = _find_header(ws, "Categoria", "Veículo")
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Header com 'Categoria' e 'Veículo' não encontrado nas primeiras 25 linhas: {path.name}"
        )

    i_veiculo    = col_index(header, "Veículo", "Veiculo")
    i_categoria  = col_index(header, "Categoria")
    i_url        = col_index(header, "URL/APP", "URL", "Url")
    i_impressoes = col_index(header, "Total de impressões", "Total de impressoes",
                              "Impressões", "Impressoes")
    i_data       = col_index(header, "Data", "Date")
    i_estado     = col_index(header, "Estado", "State", "UF")

    if i_veiculo is None or i_impressoes is None or i_categoria is None:
        wb.close()
        raise ValueError(
            f"Colunas obrigatórias ausentes (Veículo/Total de impressões/Categoria): {path.name}"
        )

    veiculos_indevidas:         dict[str, dict] = defaultdict(dict)
    veiculos_indevidas_sem_url: dict[str, dict] = defaultdict(dict)
    veiculos_entregue:  dict[str, int]  = defaultdict(int)
    veiculos_total:     dict[str, int]  = defaultdict(int)
    MAX_POOL = 10000
    url_pool: list[dict] = []
    pool_count = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row = list(row)
        if all(v is None for v in row):
            continue
        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        veiculo   = str(row[i_veiculo]).strip()   if i_veiculo   < len(row) and row[i_veiculo]   else None
        categoria = str(row[i_categoria]).strip()  if i_categoria < len(row) and row[i_categoria] else None
        url       = str(row[i_url]).strip()        if i_url is not None and i_url < len(row) and row[i_url] else None
        impressoes = to_int(row[i_impressoes] if i_impressoes < len(row) else None)

        if not veiculo or not categoria:
            continue

        if praca and i_estado is not None and i_estado < len(row):
            estado_row = str(row[i_estado]).strip().upper() if row[i_estado] else ""
            if estado_row and estado_row != praca.upper().strip():
                veiculos_total[veiculo] += impressoes
                continue

        cat_str = categoria.strip()
        veiculos_indevidas[veiculo][cat_str] = (
            veiculos_indevidas[veiculo].get(cat_str, 0) + impressoes
        )
        if not url:
            veiculos_indevidas_sem_url[veiculo][cat_str] = (
                veiculos_indevidas_sem_url[veiculo].get(cat_str, 0) + impressoes
            )

        veiculos_entregue[veiculo] += impressoes
        veiculos_total[veiculo] += impressoes

        if url and cat_str:
            pool_count += 1
            entry = {"url": url, "categoria": categoria, "veiculo": veiculo, "impressoes": impressoes}
            if len(url_pool) < MAX_POOL:
                url_pool.append(entry)
            else:
                idx = random.randint(0, pool_count - 1)
                if idx < MAX_POOL:
                    url_pool[idx] = entry

    wb.close()

    results: list[dict] = []
    for veiculo in veiculos_entregue:
        results.append({
            "veiculo":           veiculo,
            "tipo_compra":       None,
            "contratado":        None,
            "entregue":          veiculos_entregue[veiculo],
            "entregue_total":    veiculos_total[veiculo],
            "cliques":           None,
            "viewables":         None,
            "viewability":       None,
            "indevidas":         dict(veiculos_indevidas[veiculo]),
            "indevidas_sem_url": dict(veiculos_indevidas_sem_url[veiculo]),
            "url_sample":        url_pool if not results else [],
            "formato_detectado": "sense_verif",
        })

    return results


# ── CLI ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parser SENSE")
    ap.add_argument("modo", choices=["comp", "verif"])
    ap.add_argument("arquivo")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    fn = parse_comprovante if args.modo == "comp" else parse_verif
    try:
        res = fn(args.arquivo, data_ini=cli_date(args.ini), data_fim=cli_date(args.fim))
        for r in res:
            disp = {k: v for k, v in r.items() if k != "url_sample"}
            disp["url_sample_count"] = len(r.get("url_sample", []))
            print(json.dumps(disp, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, ensure_ascii=False))
        sys.exit(1)
