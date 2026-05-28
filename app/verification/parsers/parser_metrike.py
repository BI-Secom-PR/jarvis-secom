"""
Parser METRIKE — comprovante de entrega + verification de URLs.

Comprovante:
  Sheet "Worksheet". Cabeçalho da tabela na linha ~12.
  Valores totais do veículo ficam ACIMA da tabela (col E=label, F=valor):
    Total Contratado, Total impressões entregue, Total clique entregue, Veículo.
  Tabela: Veículo, Data, Campanha, ID, Placement ID, Formato, Dimensão,
          Segmentação, Descrição, Linha criativa, Contratado, Impressões,
          Cliques, Viewable, VA%, Tipo de compra.

Verification:
  Sheet única (índice 0). Cabeçalho na linha ~5.
  Colunas: Data, Veículo, Impressões, Placement, Zone, Formato, Categoria, Url.
"""

import json
import random
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from category_map import INDEVIDAS_ZERO, normaliza_categoria
from parser_utils import col_index, parse_date, to_int, cli_date, vehicle_from_filename


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _find_header(ws):
    veiculo_variants  = {"veículo", "veiculo", "vehicle"}
    entregue_variants = {"impressões", "impressoes", "impressions",
                         "entregues", "entregue",
                         "views", "view"}
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if lows & veiculo_variants and lows & entregue_variants:
            return i, vals
    return None, []


_VEICULO_VARIANTS = {"veículo", "veiculo", "vehicle", "veículos", "veiculos"}

def _find_verif_header(ws):
    """Detecta linha de cabeçalho do arquivo de verification.
    Requer 'categoria' + algum variant de 'veículo'. URL é opcional.
    """
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if "categoria" in lows and lows & _VEICULO_VARIANTS:
            return i, vals
    return None, []


# ── parse_comprovante ────────────────────────────────────────────────────────────

def _parse_header_meta(ws) -> dict:
    """Lê valores do cabeçalho acima da tabela (col E=label, F=valor).
    
    Retorna dict com chaves: contratado, entregue, cliques, veiculo.
    """
    meta: dict = {}
    for row in ws.iter_rows(max_row=10, values_only=True):
        vals = list(row)
        label = str(vals[4]).strip() if len(vals) > 4 and vals[4] else ""
        value = vals[5] if len(vals) > 5 else None
        if "Total Contratado" in label:
            meta["contratado"] = to_int(value) or 0
        elif "impressões" in label.lower() or "impressoes" in label.lower():
            meta["entregue"] = to_int(value) or 0
        elif "clique" in label.lower():
            meta["cliques"] = to_int(value)
        elif label in ("Veículo:", "Veiculo:"):
            meta["veiculo"] = str(value).strip() if value else None
    return meta


def parse_comprovante(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia comprovante METRIKE usando valores do cabeçalho acima da tabela.
    
    Os totais (contratado, entregue, cliques, veiculo) vêm das linhas 5-8.
    As linhas individuais da tabela são usadas apenas para viewability.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # ── Valores do cabeçalho (acima da tabela) ──────────────────────────────
    meta = _parse_header_meta(ws)
    veiculo_nome = meta.get("veiculo") or vehicle_from_filename(filepath)
    contratado = meta.get("contratado")
    entregue_total = meta.get("entregue")
    cliques_total = meta.get("cliques")

    # ── Encontrar header da tabela ──────────────────────────────────────────
    header_row_idx, header = _find_header(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Cabeçalho com 'Veículo' e 'Impressões'/'Views' não encontrado: {path.name}"
        )

    i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions",
                              "Entregues", "Entregue", "Views", "View")
    i_viewable   = col_index(header, "Viewable", "Viewables", "Vieweables")
    i_data       = col_index(header, "Data", "Date")

    if i_impressoes is None:
        wb.close()
        raise ValueError(f"Coluna obrigatória ausente (Impressões): {path.name}")

    # ── Linhas individuais → apenas viewability ─────────────────────────────
    total_impressions = 0
    total_viewables   = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        # Pula linhas #TOTAL, "Total por placement_id" etc.
        veiculo_raw = row[0] if row[0] is not None else ""
        data_raw = row[i_data] if i_data is not None and i_data < len(row) else None
        data_str = str(data_raw).strip() if data_raw is not None else ""

        vname = str(veiculo_raw).strip() if veiculo_raw else ""
        if "#TOTAL" in vname.upper() or "#TOTAL" in data_str.upper():
            continue
        if data_raw is not None:
            d = parse_date(data_raw)
            if d is None:
                continue
            if data_ini and d < data_ini:
                continue
            if data_fim and d > data_fim:
                continue

        imp = to_int(row[i_impressoes] if i_impressoes < len(row) else None)
        if imp == 0:
            continue

        total_impressions += imp
        if i_viewable is not None and i_viewable < len(row):
            total_viewables += to_int(row[i_viewable])

    wb.close()

    # Se o header não tinha entregue, usa a soma das linhas individuais
    if entregue_total is None:
        entregue_total = total_impressions

    viewability = (
        round(total_viewables / entregue_total * 100, 2)
        if total_viewables > 0 and entregue_total > 0
        else None
    )

    return [{
        "veiculo":           veiculo_nome,
        "tipo_compra":       None,
        "contratado":        contratado,
        "entregue":          entregue_total,
        "cliques":           cliques_total,
        "viewables":         total_viewables or None,
        "viewability":       viewability,
        "indevidas":         {},
        "url_sample":        [],
        "formato_detectado": "metrike_comprovante",
    }]


# ── parse_verif ──────────────────────────────────────────────────────────────────

def parse_verif(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
    praca: str | None = None,
) -> list[dict]:
    """Parseia verification METRIKE (Data, Veículo, Categoria, Url).

    praca — sigla do estado (ex.: 'SP') para filtrar pela coluna Estado.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row_idx, header = _find_verif_header(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Header com 'Categoria' e 'Veículo' não encontrado nas primeiras 25 linhas: {path.name}"
        )

    i_veiculo    = col_index(header, "Veículo", "Veiculo", "Vehicle",
                              "Veículos", "Veiculos")
    i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions")
    i_categoria  = col_index(header, "Categoria", "Category")
    i_url        = col_index(header, "Url", "URL", "url")
    i_data       = col_index(header, "Data", "Date")
    i_estado     = col_index(header, "Estado", "State", "UF")

    if i_veiculo is None or i_impressoes is None or i_categoria is None:
        wb.close()
        raise ValueError(
            f"Colunas obrigatórias ausentes (Veículo/Impressões/Categoria): {path.name}"
        )

    veiculos_indevidas: dict[str, dict] = defaultdict(lambda: dict(INDEVIDAS_ZERO))
    veiculos_entregue:  dict[str, int]  = defaultdict(int)
    veiculos_total:     dict[str, int]  = defaultdict(int)
    MAX_POOL = 10000
    url_pool: list[dict] = []
    pool_count = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue
        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        veiculo   = str(row[i_veiculo]).strip()   if i_veiculo   < len(row) and row[i_veiculo]   else None
        categoria = str(row[i_categoria]).strip()  if i_categoria < len(row) and row[i_categoria] else None
        url       = str(row[i_url]).strip()        if i_url is not None and i_url < len(row) and row[i_url] else None
        impressoes = to_int(row[i_impressoes] if i_impressoes < len(row) else None)

        if not veiculo or not categoria:
            continue

        if praca and i_estado is not None and i_estado < len(row):
            estado_row = str(row[i_estado]).strip().upper() if row[i_estado] else ""
            if estado_row and estado_row != praca.upper().strip():
                # Mesmo filtrando a linha, somamos ao total geral do veículo para o DIF
                veiculos_total[veiculo] += impressoes
                continue

        cat_key = normaliza_categoria(categoria)
        if cat_key:
            veiculos_indevidas[veiculo][cat_key] = (
                veiculos_indevidas[veiculo].get(cat_key, 0) + impressoes
            )

        veiculos_entregue[veiculo] += impressoes
        veiculos_total[veiculo] += impressoes

        if url and cat_key:
            pool_count += 1
            entry = {"url": url, "categoria": categoria, "veiculo": veiculo, "impressoes": impressoes}
            if len(url_pool) < MAX_POOL:
                url_pool.append(entry)
            else:
                idx = random.randint(0, pool_count - 1)
                if idx < MAX_POOL:
                    url_pool[idx] = entry

    wb.close()

    results: list[dict] = []
    for veiculo in veiculos_entregue:
        results.append({
            "veiculo":           veiculo,
            "tipo_compra":       None,
            "contratado":        None,
            "entregue":          veiculos_entregue[veiculo],
            "entregue_total":    veiculos_total[veiculo],
            "cliques":           None,
            "viewables":         None,
            "viewability":       None,
            "indevidas":         dict(veiculos_indevidas[veiculo]),
            "url_sample":        url_pool if not results else [],
            "formato_detectado": "metrike_verif",
        })

    return results


# ── CLI ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parser METRIKE")
    ap.add_argument("modo", choices=["comp", "verif"])
    ap.add_argument("arquivo")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    fn = parse_comprovante if args.modo == "comp" else parse_verif
    try:
        res = fn(args.arquivo, data_ini=cli_date(args.ini), data_fim=cli_date(args.fim))
        for r in res:
            disp = {k: v for k, v in r.items() if k != "url_sample"}
            disp["url_sample_count"] = len(r.get("url_sample", []))
            print(json.dumps(disp, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, ensure_ascii=False))
        sys.exit(1)
