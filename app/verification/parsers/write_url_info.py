"""
Escreve o levantamento de URLs indevidas (análise IA) na coluna 30 (URL info)
do consolidado verificado.

Uso:
  python3 write_url_info.py <xlsx_path> '<json>'

Onde <json> é um objeto {veiculo: "linha1\nlinha2\n", ...} mapeando o nome
do veículo conforme aparece na coluna A do consolidado para o texto a inserir.
"""

import sys
import json
import openpyxl

COL_VEICULO  = 1
COL_URL_INFO = 30
DATA_START   = 9


def main():
    if len(sys.argv) < 3:
        print("Uso: write_url_info.py <xlsx> '<json>'", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    url_info: dict[str, str] = json.loads(sys.argv[2])

    if not url_info:
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    for row_idx in range(DATA_START, (ws.max_row or 0) + 1):
        cell_val = ws.cell(row=row_idx, column=COL_VEICULO).value
        if cell_val is None:
            continue
        vname = str(cell_val).strip()
        if vname in url_info:
            ws.cell(row=row_idx, column=COL_URL_INFO).value = url_info[vname].rstrip("\n")

    wb.save(xlsx_path)
    wb.close()


if __name__ == "__main__":
    main()
