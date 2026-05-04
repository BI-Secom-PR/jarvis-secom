"""
Parser ADFORCE — comprovante de entrega + verification de URLs.

Comprovante:
  Sheet única ("sheet1"). Cabeçalho na linha ~7 com Veículo + Entregues.
  Não há linhas #TOTAL: viewability calculada por média ponderada das linhas
  de dados (coluna "Viewability" em decimal, ex.: 0.622 → 62.2%).

Verification:
  Múltiplas sheets; sheets com nome começando por "ABAT" são resumos — pular.
  Cada sheet de dados tem: Veículo, Categoria, Url.
  Categorias podem ser compostas, ex.: "Acidentes, violência, crime".
"""

import io
import json
import random
import re
import sys
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from category_map import INDEVIDAS_ZERO, normaliza_categoria
from parser_utils import col_index, parse_date, to_float, to_int, cli_date, vehicle_from_filename


# ── Helpers ─────────────────────────────────────────────────────────────────────

_NAN_CELL_RE = re.compile(rb"<v>NaN</v>", re.IGNORECASE)


def _load_workbook_safe(path: str, read_only: bool = False) -> openpyxl.Workbook:
    """Loads workbook, sanitizing NaN cell values that crash openpyxl."""
    try:
        return openpyxl.load_workbook(path, read_only=read_only, data_only=True)
    except Exception:
        pass
    # Sanitize <v>NaN</v> in worksheet XML, then retry
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/"):
                    data = _NAN_CELL_RE.sub(b"", data)
                zout.writestr(item, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf, read_only=read_only, data_only=True)


def _normalize_va(v) -> float | None:
    """Decimal (0.622) → percentagem (62.2). Mantém se já > 1."""
    f = to_float(v)
    if f is None:
        return None
    return round(f * 100, 2) if f <= 1.0 else round(f, 2)


def _find_header(ws):
    """Varre até 25 linhas buscando Impressões/Entregues com pelo menos uma coluna âncora."""
    entregue_variants = {"impressões", "impressoes", "impressions",
                         "entregues", "entregue", "entregues/impressões",
                         "entregues/impressoes", "views"}
    # Colunas âncora: confirmam que é um header de tabela, não um label de metadado
    anchor_variants = {"veículo", "veiculo", "vehicle", "placement", "data", "date", "formato", "canal"}
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if (lows & entregue_variants) and (lows & anchor_variants):
            return i, vals
    return None, []


def _find_verif_header(ws):
    """Varre até 25 linhas buscando Categoria + coluna contendo 'url'."""
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        lows = {v.lower() for v in vals}
        if "categoria" in lows and lows & {"veículo", "veiculo", "vehicle", "veículos", "veiculos"}:
            return i, vals
    return None, []


SKIP_VEHICLE_NAMES = {"---", "-", "grand total:", "grand total", "total"}


# ── parse_comprovante ────────────────────────────────────────────────────────────

def parse_comprovante(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """Parseia comprovante ADFORCE (sheet única, viewability por média ponderada)."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = _load_workbook_safe(str(path))
    ws = wb.worksheets[0]

    header_row_idx, header = _find_header(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Cabeçalho com coluna de Impressões/Entregues não encontrado: {path.name}"
        )

    i_veiculo    = col_index(header, "Veículo", "Veiculo", "Vehicle")
    # Prefer "Impressões" over "Entregues": for CPV files both columns exist and
    # "Impressões" is the ad displays (= consolidado col 5), while "Entregues" = plays.
    i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions",
                              "Entregues/Impressões", "Entregues/Impressoes")
    if i_impressoes is None:
        i_impressoes = col_index(header, "Entregues", "Entregue")
    i_cliques    = col_index(header, "Cliques", "Clicks", "Cliques Únicos")
    i_views50    = col_index(header, "100%")
    i_viewable   = col_index(header, "Viewable", "Viewables")
    i_viewability= col_index(header, "VA%", "Viewability", "VA %", "View%",
                              "Viewability (IAB)", "VA (IAB)")
    i_contratado = col_index(header, "Contratado", "Contracted")
    i_data       = col_index(header, "Data", "Date")

    if i_impressoes is None:
        wb.close()
        raise ValueError(f"Coluna de Impressões/Entregues não encontrada: {path.name}")

    # Arquivo de veículo único: infere nome do veículo pelo nome do arquivo.
    single_vehicle = vehicle_from_filename(filepath)

    entregue:   dict[str, int]   = defaultdict(int)
    cliques:    dict[str, int]   = defaultdict(int)
    views50:    dict[str, int]   = defaultdict(int)
    viewables:  dict[str, int]   = defaultdict(int)
    contratado: dict[str, int]   = {}
    viewability:dict[str, float] = {}
    va_wsum:    dict[str, float] = defaultdict(float)
    va_weight:  dict[str, int]   = defaultdict(int)

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        if i_veiculo is not None:
            veiculo_raw = row[i_veiculo] if i_veiculo < len(row) else None
            vname = str(veiculo_raw).strip() if veiculo_raw is not None else ""
            if not vname:
                vname = single_vehicle
            elif vname.lower() in SKIP_VEHICLE_NAMES:
                continue
            if vname.endswith(" Total") or vname.endswith(" total"):
                continue
        else:
            vname = single_vehicle

        if i_data is not None and i_data < len(row):
            raw_date = row[i_data]
            d = parse_date(raw_date)
            if d is None:
                continue  # linha de total/subtotal ou data vazia (evita dupla contagem)
            if data_ini and d < data_ini:
                continue
            if data_fim and d > data_fim:
                continue

        imp      = to_int(row[i_impressoes] if i_impressoes < len(row) else None)
        cliq_val = to_int(row[i_cliques]   if i_cliques  is not None and i_cliques  < len(row) else None)
        v50_val  = to_int(row[i_views50]   if i_views50  is not None and i_views50  < len(row) else None)
        va_val   = to_int(row[i_viewable]  if i_viewable is not None and i_viewable < len(row) else None)

        if imp == 0 and cliq_val == 0 and v50_val == 0 and va_val == 0:
            continue

        if imp:
            entregue[vname] += imp
        cliques[vname]  += cliq_val
        views50[vname]  += v50_val
        viewables[vname] += va_val
        if i_viewability is not None and i_viewability < len(row) and imp > 0:
            va = _normalize_va(row[i_viewability])
            if va is not None:
                va_wsum[vname]   += va * imp
                va_weight[vname] += imp

        if i_contratado is not None and i_contratado < len(row):
            c = to_int(row[i_contratado])
            if c > 0 and vname not in contratado:
                contratado[vname] = c

    wb.close()

    if not entregue:
        return []

    # Viewability: média ponderada (sem linhas #TOTAL no ADFORCE)
    for vname in entregue:
        if vname not in viewability and va_weight[vname] > 0:
            viewability[vname] = round(va_wsum[vname] / va_weight[vname], 2)

    return [
        {
            "veiculo":           vname,
            "tipo_compra":       None,
            "contratado":        contratado.get(vname),
            "entregue":          imp,
            "cliques":           cliques[vname] or None,
            "views":             views50[vname] or None,
            "viewables":         viewables[vname] or None,
            "viewability":       viewability.get(vname),
            "indevidas":         {},
            "url_sample":        [],
            "formato_detectado": "adforce_comprovante",
        }
        for vname, imp in entregue.items()
    ]


# ── parse_verif ──────────────────────────────────────────────────────────────────

def _get_verif_sheets(wb):
    """Retorna sheets de dados (pula sheets ABAT que são resumos)."""
    return [
        ws for ws in wb.worksheets
        if not ws.title.strip().upper().startswith("ABAT")
    ] or list(wb.worksheets)


def _is_flat_format(wb) -> bool:
    """
    Detecta o formato flat ADFORCE (sheet única "Result 1" com campaignUuid na col A).
    Este formato tem uma linha de cabeçalho na linha 1 sem metadados acima.
    """
    ws = wb.worksheets[0]
    first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
    if first_row is None:
        return False
    vals = {str(v).strip().lower() for v in first_row if v is not None}
    return "campaignuuid" in vals


def _parse_verif_flat(wb, data_ini, data_fim) -> tuple[dict, dict, list, int, dict, dict]:
    """
    Parseia o formato flat ADFORCE (Result 1, header na linha 1).
    Indevidas = soma de CPM+CPC+CPV+CPCV por categoria (apenas uma terá valor por linha).
    Retorna (indev, entregue_dict, url_pool, pool_count, cpv_indev, cpv_total_by_vehicle).
    cpv_indev é separado para indevidas; cpv_total_by_vehicle alimenta DIF views.
    """
    ws = wb.worksheets[0]
    header = [str(v).strip() if v is not None else "" for v in
              next(ws.iter_rows(max_row=1, values_only=True))]

    i_vehicle    = col_index(header, "vehicle", "Vehicle", "Veículo", "Veiculo")
    i_categories = col_index(header, "categories", "Categories", "Categoria", "Category")
    i_url        = col_index(header, "url", "URL", "Url")
    i_cpm        = col_index(header, "cpm", "CPM", "Impressoes", "Impressões", "Impressions")
    i_cpc        = col_index(header, "cpc", "CPC", "Cliques", "Clicks")
    i_cpv        = col_index(header, "cpv", "CPV", "Views", "Visualizações", "Visualizacoes")
    i_cpcv       = col_index(header, "cpcv", "CPCV")
    i_data       = col_index(header, "date", "Date", "Data")

    if i_vehicle is None or i_categories is None:
        raise ValueError("Colunas 'vehicle'/'categories' não encontradas no formato flat ADFORCE")

    indev:     dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    cpv_indev: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    cpv_total_by_vehicle: dict[str, int] = defaultdict(int)
    veiculos_entregue: dict[str, int] = defaultdict(int)
    MAX_POOL = 500
    url_pool: list[dict] = []
    pool_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        if (data_ini or data_fim) and i_data is not None and i_data < len(row):
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        veiculo   = str(row[i_vehicle]).strip()    if i_vehicle    < len(row) and row[i_vehicle]    else None
        categoria = str(row[i_categories]).strip() if i_categories < len(row) and row[i_categories] else None
        url       = str(row[i_url]).strip()        if i_url is not None and i_url < len(row) and row[i_url] else None

        if not veiculo:
            continue

        def _metric(i_col) -> int:
            if i_col is None or i_col >= len(row):
                return 0
            return to_int(row[i_col]) or 0

        v_cpm  = _metric(i_cpm)
        v_cpc  = _metric(i_cpc)
        v_cpv  = _metric(i_cpv)
        v_cpcv = _metric(i_cpcv)
        total_val = v_cpm + v_cpc + v_cpv + v_cpcv

        veiculos_entregue[veiculo] += total_val
        cpv_total_by_vehicle[veiculo] += v_cpv

        if not categoria:
            continue

        cat_key = normaliza_categoria(categoria)
        if not cat_key:
            continue

        if total_val > 0:
            indev[veiculo][cat_key] += total_val
        if v_cpv > 0:
            cpv_indev[veiculo][cat_key] += v_cpv

        if url and total_val > 0:
            pool_count += 1
            entry = {"url": url, "categoria": categoria, "veiculo": veiculo,
                     "impressoes": total_val}
            if len(url_pool) < MAX_POOL:
                url_pool.append(entry)
            else:
                idx = random.randint(0, pool_count - 1)
                if idx < MAX_POOL:
                    url_pool[idx] = entry

    return indev, veiculos_entregue, url_pool, pool_count, cpv_indev, cpv_total_by_vehicle


def _parse_verif_multitab(wb, data_ini, data_fim) -> tuple[dict, dict, list, int, dict]:
    """
    Parseia o formato multi-sheet ADFORCE (uma sheet por veículo, pula ABAT).
    Indevidas = coluna Impressões apenas (as demais colunas não correspondem a indevidas).
    Retorna (indev, entregue_dict, url_pool, pool_count, cpv_indev).
    cpv_indev é vazio neste formato (Impressões não distingue tipo de métrica).
    """
    sheets = _get_verif_sheets(wb)
    indev: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    veiculos_entregue: dict[str, int] = defaultdict(int)
    MAX_POOL = 500
    url_pool: list[dict] = []
    pool_count = 0
    found_header = False

    for ws in sheets:
        header_row_idx, header = _find_verif_header(ws)
        if header_row_idx is None:
            continue
        found_header = True

        i_veiculo   = col_index(header, "Veículo", "Veiculo", "Vehicle",
                                 "Veículos", "Veiculos")
        i_categoria = col_index(header, "Categoria", "Category")
        i_url       = col_index(header, "Url", "URL", "url", "URL Veiculada",
                                 "Url Veiculada")
        i_data      = col_index(header, "Data", "Date")

        # Metric columns: each section has Impressões/Cliques/Visualizações/Vis.Completas
        i_imp      = col_index(header, "Impressões", "Impressoes", "Impressions",
                                "Impressões Totais", "Impressoes Totais")
        i_cliques  = col_index(header, "Cliques", "Clicks")
        i_vis      = col_index(header, "Visualizações", "Visualizacoes", "Views",
                                "Visualizações CPV", "Visualizacoes CPV")
        i_vis_comp = col_index(header, "Visualizações Completas",
                                "Visualizacoes Completas", "Complete Views",
                                "Completed Views", "Visualizações Completadas")

        if i_veiculo is None or i_categoria is None:
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(v is None for v in row):
                continue
            if (data_ini or data_fim) and i_data is not None and i_data < len(row):
                d = parse_date(row[i_data])
                if d:
                    if data_ini and d < data_ini:
                        continue
                    if data_fim and d > data_fim:
                        continue

            veiculo   = str(row[i_veiculo]).strip()   if i_veiculo   < len(row) and row[i_veiculo]   else None
            categoria = str(row[i_categoria]).strip()  if i_categoria < len(row) and row[i_categoria] else None
            url       = str(row[i_url]).strip()        if i_url is not None and i_url < len(row) and row[i_url] else None

            if not veiculo or not categoria:
                continue

            def _metric(i_col) -> int:
                if i_col is None or i_col >= len(row):
                    return 0
                return to_int(row[i_col]) or 0

            v_imp = _metric(i_imp)

            veiculos_entregue[veiculo] += v_imp

            cat_key = normaliza_categoria(categoria)
            if not cat_key:
                continue

            if v_imp > 0:
                indev[veiculo][cat_key] += v_imp

            if url and v_imp > 0:
                pool_count += 1
                entry = {"url": url, "categoria": categoria, "veiculo": veiculo,
                         "impressoes": v_imp}
                if len(url_pool) < MAX_POOL:
                    url_pool.append(entry)
                else:
                    idx = random.randint(0, pool_count - 1)
                    if idx < MAX_POOL:
                        url_pool[idx] = entry

    if not found_header:
        raise ValueError(
            "Header com 'Categoria' e 'Veículo' não encontrado em nenhuma sheet"
        )

    return indev, veiculos_entregue, url_pool, pool_count, {}


def parse_verif(
    filepath: str,
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """
    Parseia verification ADFORCE — detecta automaticamente o formato:
      - Flat (single sheet "Result 1" com campaignUuid): FEVEREIRO/MARÇO style
      - Multi-sheet (uma tab por veículo, pula ABAT): JANEIRO style
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = _load_workbook_safe(str(path), read_only=True)

    if _is_flat_format(wb):
        formato = "adforce_verif_flat"
        indev, veiculos_entregue, url_pool, _, cpv_indev, cpv_total = _parse_verif_flat(
            wb, data_ini, data_fim
        )
    else:
        formato = "adforce_verif_multitab"
        indev, veiculos_entregue, url_pool, _, cpv_indev = _parse_verif_multitab(
            wb, data_ini, data_fim
        )
        cpv_total = {}

    wb.close()

    results: list[dict] = []
    for veiculo in veiculos_entregue:
        cpv_total_vehicle = cpv_total.get(veiculo) if formato == "adforce_verif_flat" else None
        if cpv_total_vehicle is None:
            cpv_total_vehicle = sum(cpv_indev.get(veiculo, {}).values())
        results.append({
            "veiculo":           veiculo,
            "contratado":        None,
            "entregue":          veiculos_entregue[veiculo],
            "views":             cpv_total_vehicle or None,
            "cliques":           None,
            "viewables":         None,
            "viewability":       None,
            "indevidas":         dict(indev.get(veiculo, {})),
            "url_sample":        url_pool if not results else [],
            "formato_detectado": formato,
        })

    return results


# ── CLI ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parser ADFORCE")
    ap.add_argument("modo", choices=["comp", "verif"])
    ap.add_argument("arquivo")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    fn = parse_comprovante if args.modo == "comp" else parse_verif
    try:
        res = fn(args.arquivo, data_ini=cli_date(args.ini), data_fim=cli_date(args.fim))
        for r in res:
            disp = {k: v for k, v in r.items() if k != "url_sample"}
            disp["url_sample_count"] = len(r.get("url_sample", []))
            print(json.dumps(disp, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, ensure_ascii=False))
        sys.exit(1)
