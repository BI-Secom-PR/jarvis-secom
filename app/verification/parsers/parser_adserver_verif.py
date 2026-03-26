"""
Parser para arquivos de Verification de Adserver (R7, Terra, UOL e similares).

Detectado por: nome de arquivo começa com 'verification' (qualquer case) sem
espaço-hífen após (distingue de VETTA 'Verification - *.xlsx').

Estrutura esperada:
  Linhas 1–N: metadados livres (Cliente, Campanha, etc.)
  Linha M:    cabeçalho com colunas Data, Veículo, Impressões, Placement,
              Zone, Formato, Categoria, Url
  Linha M+1+: dados — cada linha = grupo de impressões em uma URL com uma categoria

Saída (list[dict] — um dict por veículo encontrado no arquivo):
  {
    veiculo:          str,
    tipo_compra:      None,         # não disponível neste formato
    contratado:       None,
    entregue:         int,          # soma de Impressões do veículo
    cliques:          None,
    viewables:        None,
    viewability:      None,
    indevidas:        {cat: int},   # 9 categorias SECOM
    url_sample:       [{url, categoria}],  # 5% aleatório do arquivo inteiro
    formato_detectado: "adserver_verif",
  }
"""

import json
import random
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from category_map import INDEVIDAS_ZERO, normaliza_categoria
from parser_utils import col_index, parse_date, to_int, cli_date


def _find_header_row(ws) -> tuple[int | None, list[str]]:
    """
    Varre as primeiras 25 linhas procurando o header com 'Categoria' + 'Url'.

    Retorna (row_idx_1based, header_values_list) ou (None, []).
    """
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        vals_lower = [v.lower() for v in vals]
        if "categoria" in vals_lower and "url" in vals_lower:
            return i, vals
    return None, []



def parse(
    filepath: str,
    verif_paths: list[str] | None = None,  # ignorado (compatibilidade de interface)
    data_ini: date | None = None,
    data_fim: date | None = None,
) -> list[dict]:
    """
    Parseia um arquivo de verification de adserver.

    Retorna list[dict] — um dict por veículo encontrado no arquivo.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    header_row_idx, header = _find_header_row(ws)
    if header_row_idx is None:
        wb.close()
        raise ValueError(
            f"Header com 'Categoria' e 'Url' não encontrado nas primeiras 25 linhas: {path.name}"
        )

    # Índices de colunas (0-based)
    i_data       = col_index(header, "Data", "Date")
    i_veiculo    = col_index(header, "Veículo", "Veiculo", "Vehicle")
    i_impressoes = col_index(header, "Impressões", "Impressoes", "Impressions")
    i_categoria  = col_index(header, "Categoria", "Category")
    i_url        = col_index(header, "Url", "URL", "url")

    if i_veiculo is None or i_impressoes is None or i_categoria is None:
        wb.close()
        raise ValueError(
            f"Colunas obrigatórias ausentes (Veículo/Impressões/Categoria): {path.name}"
        )

    # Acumuladores por veículo
    veiculos_indevidas: dict[str, dict[str, int]] = defaultdict(
        lambda: dict(INDEVIDAS_ZERO)
    )
    veiculos_entregue: dict[str, int] = defaultdict(int)

    # Pool de URLs para amostragem — reservoir sampling, máx 500 entradas
    MAX_POOL = 500
    url_pool: list[dict] = []
    pool_count = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        # Filtro de data
        if (data_ini or data_fim) and i_data is not None:
            d = parse_date(row[i_data])
            if d:
                if data_ini and d < data_ini:
                    continue
                if data_fim and d > data_fim:
                    continue

        veiculo   = str(row[i_veiculo]).strip() if row[i_veiculo] else None
        categoria = str(row[i_categoria]).strip() if row[i_categoria] else None
        url       = str(row[i_url]).strip() if i_url is not None and row[i_url] else None
        impressoes = to_int(row[i_impressoes])

        if not veiculo or not categoria:
            continue

        # Mapear categoria → chave SECOM
        cat_key = normaliza_categoria(categoria)
        if cat_key:
            veiculos_indevidas[veiculo][cat_key] = (
                veiculos_indevidas[veiculo].get(cat_key, 0) + impressoes
            )

        veiculos_entregue[veiculo] += impressoes

        # Apenas URLs de categorias indevidas vão para a amostra de IA
        if url and cat_key:
            pool_count += 1
            entry = {"url": url, "categoria": categoria, "veiculo": veiculo}
            if len(url_pool) < MAX_POOL:
                url_pool.append(entry)
            else:
                # Reservoir sampling: replace a random earlier entry
                idx = random.randint(0, pool_count - 1)
                if idx < MAX_POOL:
                    url_pool[idx] = entry

    wb.close()

    # Amostra de 5% do pool de URLs (calculada uma vez, só no primeiro resultado)
    sample_size = max(1, len(url_pool) // 20) if url_pool else 0
    url_sample = (
        random.sample(url_pool, min(sample_size, len(url_pool)))
        if url_pool
        else []
    )

    # Montar resultados (um dict por veículo)
    results: list[dict] = []
    for veiculo in veiculos_entregue:
        results.append({
            "veiculo":          veiculo,
            "tipo_compra":      None,
            "contratado":       None,
            "entregue":         veiculos_entregue[veiculo],
            "cliques":          None,
            "viewables":        None,
            "viewability":      None,
            "indevidas":        dict(veiculos_indevidas[veiculo]),
            "url_sample":       url_sample if not results else [],  # só no primeiro
            "formato_detectado": "adserver_verif",
        })

    return results


# ── CLI para testes ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description="Parser de verification de adserver (R7/Terra/UOL)"
    )
    ap.add_argument("arquivo", help="Arquivo verification_*.xlsx")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    args = ap.parse_args()

    try:
        resultados = parse(
            args.arquivo,
            data_ini=cli_date(args.ini),
            data_fim=cli_date(args.fim),
        )
        # Suprimir url_sample no print para não poluir o terminal
        for r in resultados:
            r_display = {k: v for k, v in r.items() if k != "url_sample"}
            r_display["url_sample_count"] = len(r.get("url_sample", []))
            print(json.dumps(r_display, indent=2, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, indent=2, ensure_ascii=False))
        sys.exit(1)
