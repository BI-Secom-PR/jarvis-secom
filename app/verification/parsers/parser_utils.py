"""
Utilitários compartilhados pelos parsers de verification/comprovante.
"""

import random
import re
from datetime import date, datetime
from pathlib import Path

import fastxlsx
import openpyxl


def load_workbook_fast(path: str, read_only: bool = True):
    """Loads an xlsx workbook via calamine (Rust, ~10-50x faster than openpyxl
    on large sheets — verification files can be 10-20MB+ and were hitting
    Vercel's 300s cap under openpyxl). Falls back to openpyxl if calamine
    can't parse the file."""
    try:
        return fastxlsx.load_workbook(path)
    except Exception:
        pass
    return openpyxl.load_workbook(path, read_only=read_only, data_only=True)


def to_int(v) -> int:
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", ".")))
    except (ValueError, TypeError):
        return 0


def to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def parse_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                pass
    # Excel serial number stored as float (openpyxl returns raw float for unformatted date cells)
    if isinstance(v, (int, float)) and 1 < float(v) < 2958466:
        try:
            from openpyxl.utils.datetime import from_excel
            result = from_excel(float(v))
            return result.date() if isinstance(result, datetime) else result
        except Exception:
            pass
    return None


def col_index(header: list[str], *names: str) -> int | None:
    """Índice (0-based) da coluna cujo nome (case-insensitive) bate com o primeiro
    candidato de `names` presente no header — a ORDEM de `names` é prioridade.
    Importa quando o header tem mais de uma coluna candidata ao mesmo tempo (ex.:
    "Total de impressões" e "Impressões válidas" na mesma planilha SENSE V1) —
    antes a prioridade era por posição da coluna, não pela ordem dos candidatos,
    o que fazia colunas de fallback "vencerem" a preferida quando vinham antes
    no header."""
    header_lower = [h.lower() for h in header]
    for name in names:
        name_lower = name.lower()
        for i, h in enumerate(header_lower):
            if h == name_lower:
                return i
    return None


def parse_comprovante_cm360(ws, formato: str) -> dict | None:
    """
    Extrai o total de UM veículo de uma planilha de comprovante CM360/DFA
    (usado por DGBRASIL — um veículo por arquivo — e TERATECH — um por aba).

    Layout: metadados nas linhas 1-6, header na linha 8, dados a partir da 9.
    As linhas de total têm "Data" = "-"; há uma por Package e, no fim, o total
    geral do veículo com "Package" = "-" (bate com "Impressoes/Unidades
    Entregues" do cabeçalho na linha 5) e o veículo escrito como "<Nome> Total".
    Pegar o primeiro "Data" = "-" pegaria só o primeiro package — ex.: TERATECH
    Claro Ads tem 3 packages (1.767.981 no 1º vs 4.836.462 no total); DGBRASIL
    Sou + Favela tem 5. Sem nenhuma linha "Package" = "-" cai na primeira linha
    de total (aba de package único).

    Duas variantes de coluna: display/CPM (14 cols) e vídeo/CPV (19 cols, com
    Video Plays/Quartis/Completions). "Views" no consolidado = Video Completions.

    Retorna None quando a planilha não tem o layout de comprovante.
    """
    header = [str(v).strip() if v is not None else "" for v in
              next(ws.iter_rows(min_row=8, max_row=8, values_only=True), [])]
    if not header:
        return None

    i_veiculo     = col_index(header, "Veiculo", "Veículo")
    i_data        = col_index(header, "Data")
    i_package     = col_index(header, "Package")
    i_contratado  = col_index(header, "Contratado")
    i_impressoes  = col_index(header, "Impressoes", "Impressões")
    i_cliques     = col_index(header, "Cliques")
    i_viewable    = col_index(header, "Active View: Viewable Impressions")
    i_viewability = col_index(header, "Active View: % Viewable Impressions")
    i_completions = col_index(header, "Video Completions")

    if i_veiculo is None or i_impressoes is None or i_data is None:
        return None

    is_cpv = i_completions is not None

    totais = [
        row for row in ws.iter_rows(min_row=9, values_only=True)
        if not all(v is None for v in row)
        and i_data < len(row) and str(row[i_data]).strip() == "-"
        and row[i_veiculo] and str(row[i_veiculo]).strip()
    ]
    geral = None
    if i_package is not None:
        geral = next(
            (r for r in totais
             if i_package < len(r) and str(r[i_package]).strip() == "-"),
            None,
        )
    row = geral or (totais[0] if totais else None)
    if row is None:
        return None

    veiculo = str(row[i_veiculo]).strip()
    if veiculo.lower().endswith(" total"):
        veiculo = veiculo[:-6].strip()

    va = to_float(row[i_viewability]) if i_viewability is not None else None

    return {
        "veiculo":           veiculo,
        "tipo_compra":       "CPV" if is_cpv else "CPM",
        "contratado":        to_int(row[i_contratado]) or None if i_contratado is not None else None,
        "entregue":          to_int(row[i_impressoes]),
        "views":             to_int(row[i_completions]) if is_cpv else None,
        "cliques":           to_int(row[i_cliques]) or None if i_cliques is not None else None,
        "viewables":         to_int(row[i_viewable]) or None if i_viewable is not None else None,
        "viewability":       round(va * 100, 2) if va is not None and va <= 1.0 else va,
        "indevidas":         {},
        "url_sample":        [],
        "formato_detectado": formato,
    }


def cli_date(s: str | None) -> date | None:
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def vehicle_from_filename(filepath: str) -> str:
    """
    Inferência simples de veículo via nome do arquivo.
    Prioriza o último segmento após " - " e remove sufixos comuns.
    """
    stem = Path(filepath).stem.strip()
    if " - " in stem:
        candidate = stem.split(" - ")[-1].strip()
    else:
        candidate = stem
    candidate = re.sub(r"\b(comprovante|verification|verificacao|relatorio)\b", "", candidate, flags=re.IGNORECASE)
    candidate = re.sub(r"\s+", " ", candidate).strip(" _-")
    return candidate or stem


class StratifiedReservoir:
    """
    Reservoir sampling estratificado para o pool de URLs.

    Um reservoir global único deixa estratos raros (ex.: as poucas linhas cpv>0
    de um veículo CPV num arquivo dominado por linhas cpm de outros veículos)
    serem afogados pelo volume. Mantendo um reservoir independente por estrato
    — (veículo, categoria, métrica) — linhas raras sobrevivem garantidamente
    enquanto estratos gigantes continuam limitados a `cap` itens.
    """

    def __init__(self, cap: int = 500):
        self.cap = cap
        self._pools: dict[tuple, list[dict]] = {}
        self._counts: dict[tuple, int] = {}

    def add(self, key: tuple, entry: dict) -> None:
        pool = self._pools.setdefault(key, [])
        self._counts[key] = self._counts.get(key, 0) + 1
        if len(pool) < self.cap:
            pool.append(entry)
        else:
            idx = random.randint(0, self._counts[key] - 1)
            if idx < self.cap:
                pool[idx] = entry

    def items(self) -> list[dict]:
        return [e for pool in self._pools.values() for e in pool]
