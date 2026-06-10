"""
Engine de comparação do Sistema de Verificação SECOM.

Fluxo:
  1. Lê o consolidado (template 29 colunas) via openpyxl
  2. Separa os arquivos em 2 grupos (pelo nome do arquivo):
       a) "comprovante" → valores de entrega (impressões, cliques, viewables)
       b) "verification" → categorias indevidas + URL por linha (adserver)
  3. Faz match fuzzy entre veículos do consolidado e resultados dos parsers
  4. Compara métricas:
       - entregue/cliques/viewables → vs comprovante
       - categorias indevidas       → vs verification
  5. Escreve devolutiva na coluna 28 (Devolutiva BI SECOM) do consolidado
  6. Salva como <nome> - Verificado.xlsx (não sobrescreve o original)

Regras de divergência:
  - Entregue: |comp − consol| / consol > 2% → flaggar
  - Cada categoria indevida: verif ≠ consol → flaggar
  - Veículo no consolidado sem comprovante E sem verification → PENDENTE
  - Arquivo sem veículo no consolidado → avisado mas não grava

Colunas do consolidado (1-indexed, template 29 colunas):
  1=Veículo  3=Tipo  4=Contratado  5=Impressões  7=Cliques  9=Views
  11=Viewables  12=Viewability
  14=Conteúdo Sensível  15–22=indevidas individuais
  28=Devolutiva BI SECOM  29=Devolutiva Agência
"""

import importlib
import sys
import json
import unicodedata
import re
import random
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, process

sys.path.insert(0, str(Path(__file__).parent / "parsers"))
from category_map import normaliza_categoria  # noqa: E402

# ── Registro de parsers por adserver ───────────────────────────────────────────
PARSER_MAP: dict[str, str] = {
    "00px":     "parser_00px",
    "adforce":  "parser_adforce",
    "admotion": "parser_admotion",
    "ahead":    "parser_ahead",
    "metrike":  "parser_metrike",
    "brz":      "parser_brz",
}

# ── Constantes ─────────────────────────────────────────────────────────────────
FUZZY_THRESHOLD  = 85     # % mínimo para aceitar match de veículo
HEADER_ROW       = 8      # linha 1-indexed do cabeçalho no template
DATA_START_ROW   = 9      # primeira linha de dados (1-indexed)

# Cores para a Devolutiva no Excel (Standard Excel Theme: Green/Red/Yellow/Gray)
COLOR_OK       = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde claro
COLOR_DIV      = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Vermelho claro
COLOR_ALERTA   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarelo/Âmbar
COLOR_PENDENTE = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Cinza claro

# Mapeamento col_index(1-based) → chave interna  [template 29 colunas]
COL_VEICULO        = 1
COL_PRACA          = 2
COL_TIPO           = 3   # Objetivo de Mídia / Tipo de Compra (CPM, CPV, CPC, CPCV…)
COL_CONTRATADO     = 4
COL_IMPRESSOES     = 5
COL_CLIQUES        = 7
COL_VIEWS          = 9
COL_VIEWABLES      = 11
COL_VIEWABILITY    = 12
COL_ENTREGAS_VAL   = 13
COL_INDEVIDAS = {
    "conteudo_sensivel":  14,   # Conteúdo Sensível (agregado)
    "acidente":           15,
    "violencia":          16,
    "lingua_estrangeira": 17,
    "pornografia":        18,
    "safeframe":          19,
    "app_movel":          20,
    "teste_tag":          21,
    "nao_classificado":   22,
}
COL_DEVOLUTIVA_BI      = 28
COL_DEVOLUTIVA_AGENCIA = 29
COL_URL_INFO           = 30


# ── Detecção dinâmica de colunas por header ────────────────────────────────────

def _detect_consolidado_cols(ws) -> dict:
    """
    Lê o header (HEADER_ROW=8) e retorna posições reais das colunas,
    permitindo que cada adserver use seu próprio layout de consolidado.
    Fallback para os valores hardcoded quando uma coluna não é encontrada.
    """
    indevidas: dict[str, int] = {}
    devolutiva_bi  = COL_DEVOLUTIVA_BI
    url_info       = COL_URL_INFO
    col_views_start:  int | None = None
    col_views_50:     int | None = None
    col_views_100:    int | None = None
    col_viewables:    int | None = None
    col_viewability:  int | None = None

    for c in range(1, 36):
        raw = _cell_value(ws, HEADER_ROW, c)
        if raw is None:
            continue
        text = str(raw).strip()
        tl = text.lower()

        key = normaliza_categoria(text)
        if key:
            indevidas[key] = c

        if "devolutiva bi" in tl or ("devolutiva" in tl and "secom" in tl):
            devolutiva_bi = c
        elif "url info" in tl:
            url_info = c
        elif tl in ("views start", "views 0%"):
            col_views_start = c
        elif tl in ("views 50%", "views50%"):
            col_views_50 = c
        elif tl in ("views 100%", "views100%"):
            col_views_100 = c
        elif tl in ("va", "viewables", "viewable impressions", "impressões viewáveis", "impressoes viewaveis"):
            col_viewables = c
        elif tl in ("va%", "viewability", "viewability%", "taxa de viewability"):
            col_viewability = c

    return {
        "indevidas":       indevidas if indevidas else dict(COL_INDEVIDAS),
        "devolutiva_bi":   devolutiva_bi,
        "url_info":        url_info,
        "col_views_start": col_views_start,
        "col_views_50":    col_views_50,
        "col_views_100":   col_views_100,
        "col_viewables":   col_viewables,
        "col_viewability": col_viewability,
    }


# ── Normalização de nomes ───────────────────────────────────────────────────────
_REMOVE_SUFFIXES = re.compile(
    r"\b(s\.?a\.?|ltda\.?|eireli|me|epp|s\/a|sa)\b", re.IGNORECASE
)
_REMOVE_PUNCT = re.compile(r"[^\w\s]")
_EXTRA_SPACES = re.compile(r"\s+")


def _normalize(name: str) -> str:
    """Uppercase, sem acentos, sem pontuação, sem sufixos empresariais."""
    if not name:
        return ""
    # Remove acentos
    nfkd = unicodedata.normalize("NFKD", str(name))
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    # Uppercase
    s = ascii_str.upper()
    # Remove sufixos
    s = _REMOVE_SUFFIXES.sub(" ", s)
    # Remove pontuação (mantém espaços)
    s = _REMOVE_PUNCT.sub(" ", s)
    # Normaliza espaços
    return _EXTRA_SPACES.sub(" ", s).strip()


# ── Leitura do consolidado ──────────────────────────────────────────────────────
def _cell_value(ws, row: int, col: int):
    """Lê valor de célula ignorando fórmulas (retorna None se fórmula sem valor)."""
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if isinstance(val, str) and val.startswith("="):
        return None
    return val


def _to_int_safe(v) -> int:
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", ".")))
    except (ValueError, TypeError):
        return 0


def _to_float_safe(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _read_consolidado(ws) -> tuple[list[dict], int]:
    """
    Lê todas as linhas de dados do consolidado.
    Retorna (lista de dicts com métricas por veículo, coluna da devolutiva BI).
    Detecta posições de colunas dinamicamente pelo header — suporta layouts
    adserver-específicos sem branches por adserver.
    """
    detected = _detect_consolidado_cols(ws)
    col_indevidas    = detected["indevidas"]
    col_dev_bi       = detected["devolutiva_bi"]
    col_vs           = detected["col_views_start"]
    col_v50          = detected["col_views_50"]
    col_v100         = detected["col_views_100"]
    col_va           = detected["col_viewables"]   or COL_VIEWABLES
    col_va_pct       = detected["col_viewability"] or COL_VIEWABILITY

    rows = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        veiculo = _cell_value(ws, row_idx, COL_VEICULO)
        if not veiculo or not str(veiculo).strip():
            continue


        entregue = _to_int_safe(_cell_value(ws, row_idx, COL_IMPRESSOES))
        views    = _to_int_safe(_cell_value(ws, row_idx, COL_VIEWS))

        indevidas = {
            cat: _to_int_safe(_cell_value(ws, row_idx, col))
            for cat, col in col_indevidas.items()
        }

        # Normaliza viewability do consolidado: Excel armazena como decimal (0.7166 = 71.66%)
        va_raw = _to_float_safe(_cell_value(ws, row_idx, col_va_pct))
        if va_raw is not None:
            viewability_val = round(va_raw * 100, 2) if va_raw <= 1.0 else round(va_raw, 2)
        else:
            viewability_val = None

        tipo_raw = _cell_value(ws, row_idx, COL_TIPO)
        tipo_compra = str(tipo_raw).strip().upper() if tipo_raw else ""

        rows.append({
            "row_idx":       row_idx,
            "veiculo":       str(veiculo).strip(),
            "tipo_compra":   tipo_compra,
            "contratado":    _to_int_safe(_cell_value(ws, row_idx, COL_CONTRATADO)),
            "entregue":      entregue,
            "views":         views,
            "views_start":   _to_int_safe(_cell_value(ws, row_idx, col_vs))   if col_vs  else None,
            "views_50":      _to_int_safe(_cell_value(ws, row_idx, col_v50))  if col_v50 else None,
            "views_100":     _to_int_safe(_cell_value(ws, row_idx, col_v100)) if col_v100 else None,
            "cliques":       _to_int_safe(_cell_value(ws, row_idx, COL_CLIQUES)),
            "viewables":     _to_int_safe(_cell_value(ws, row_idx, col_va)),
            "viewability":   viewability_val,
            "indevidas":     indevidas,
        })

    return rows, col_dev_bi


def _add_optional(a: int | None, b: int | None) -> int | None:
    """Soma dois valores opcionais; retorna None somente se ambos forem None."""
    if a is None and b is None:
        return None
    return (a or 0) + (b or 0)


def _fuzzy_match(consol_norm: str, name_list: list[str], norm_dict: dict) -> tuple[dict | None, int]:
    """Retorna (result_dict, score) do melhor match fuzzy, ou (None, 0) se abaixo do threshold."""
    if not name_list:
        return None, 0
    # Match exato tem prioridade absoluta — evita que "UOL" e "UOL VAST" se confundam,
    # já que token_set_ratio trata subconjuntos de tokens como 100%.
    if consol_norm in norm_dict:
        return norm_dict[consol_norm], 100
    best = process.extractOne(
        consol_norm, name_list,
        scorer=fuzz.token_set_ratio,
        score_cutoff=FUZZY_THRESHOLD,
    )
    if best:
        return norm_dict[best[0]], best[1]
    return None, 0


def _merge_by_veiculo(results: list[dict]) -> dict[str, dict]:
    """
    Agrupa e soma resultados de múltiplos arquivos do mesmo veículo.
    Retorna dict: {nome_normalizado: dict_merged}
    """
    merged: dict[str, dict] = {}
    for r in results:
        key = _normalize(r["veiculo"])
        if key not in merged:
            merged[key] = dict(r)
            merged[key]["indevidas"]     = dict(r.get("indevidas", {}))
            merged[key]["indevidas_cpv"] = dict(r.get("indevidas_cpv", {}))
        else:
            m = merged[key]
            m["entregue"]    = (m.get("entregue") or 0) + (r.get("entregue") or 0)
            m["cliques"]     = _add_optional(m.get("cliques"),      r.get("cliques"))
            m["viewables"]   = _add_optional(m.get("viewables"),    r.get("viewables"))
            m["views_start"] = _add_optional(m.get("views_start"),  r.get("views_start"))
            m["views_50"]    = _add_optional(m.get("views_50"),     r.get("views_50"))
            m["views_100"]   = _add_optional(m.get("views_100"),    r.get("views_100"))
            for cat, val in r.get("indevidas", {}).items():
                m["indevidas"][cat] = m["indevidas"].get(cat, 0) + val
            for cat, val in r.get("indevidas_cpv", {}).items():
                m["indevidas_cpv"][cat] = m["indevidas_cpv"].get(cat, 0) + val
    return merged


# ── Comparação de métricas ──────────────────────────────────────────────────────
def _fmt_num(n: int | float) -> str:
    return f"{int(n):,}".replace(",", ".")


def _compare(
    consol_row: dict,
    comp_result: dict | None,
    verif_result: dict | None,
    verif_result_dif: dict | None = None,
) -> tuple[str, list[str]]:
    """
    Compara métricas de uma linha do consolidado com comprovante e verification.
    Retorna todas as linhas de devolutiva (OK e DIV) para rastreabilidade completa.

    Retorna: (status, linhas_devolutiva)
    status: "OK" | "DIVERGENCIA"
    """
    tem_divergencia = False
    linhas: list[str] = []

    # ── Campos do comprovante ─────────────────────────────────────────────────
    if comp_result is not None:
        # Usar breakdown granular (views_start/50/100) quando o consolidado tiver essas colunas;
        # caso contrário, cair de volta no campo legado "views".
        _VIEW_BREAKDOWN = ("views_start", "views_50", "views_100")
        has_breakdown = any(consol_row.get(k) for k in _VIEW_BREAKDOWN)
        campos_views = (
            [k for k in _VIEW_BREAKDOWN if consol_row.get(k) or comp_result.get(k)]
            if has_breakdown else ["views"]
        )

        for campo in ("entregue", "cliques", *campos_views, "viewables"):
            comp_val   = comp_result.get(campo) or 0
            consol_val = consol_row.get(campo)  or 0
            if comp_val == 0:
                continue  # sem dado no comprovante — não gera falso DIV
            if comp_val != consol_val:
                linhas.append(
                    f"DIV {campo}: comprovante {_fmt_num(comp_val)} / "
                    f"consolidado {_fmt_num(consol_val)}"
                )
                tem_divergencia = True
            else:
                linhas.append(f"OK {campo}: {_fmt_num(comp_val)}")

        comp_va   = comp_result.get("viewability")
        consol_va = consol_row.get("viewability")
        if comp_va is not None:
            if consol_va is not None and round(comp_va, 2) != round(consol_va, 2):
                linhas.append(
                    f"DIV viewability: comprovante {comp_va:.2f}% / "
                    f"consolidado {consol_va:.2f}%"
                )
                tem_divergencia = True
            else:
                ref = f"{consol_va:.2f}%" if consol_va is not None else "—"
                linhas.append(f"OK viewability: {comp_va:.2f}% (consolidado {ref})")
    else:
        linhas.append("? comprovante: veiculo nao localizado")

    # ── Indevidas ──────────────────────────────────────────────────────────────
    consol_indev = consol_row.get("indevidas", {})

    if verif_result is not None:
        # Para campanhas CPV, indevidas = coluna cpv (views); demais usam total combinado.
        tipo_compra = consol_row.get("tipo_compra", "")
        if tipo_compra == "CPV" and verif_result.get("indevidas_cpv"):
            verif_indev_raw = verif_result.get("indevidas_cpv", {})
        else:
            verif_indev_raw = verif_result.get("indevidas", {})
        verif_indev: dict[str, int] = {}
        verif_extras: dict[str, int] = {}
        for raw_cat, count in verif_indev_raw.items():
            cat_key = normaliza_categoria(raw_cat)
            if cat_key:
                verif_indev[cat_key] = verif_indev.get(cat_key, 0) + count
            else:
                verif_extras[raw_cat] = verif_extras.get(raw_cat, 0) + count

        indev_linhas: list[str] = []
        alguma_indevida_com_dado = False

        for cat, consol_val in consol_indev.items():
            comp_val = verif_indev.get(cat, 0)
            if comp_val == 0 and consol_val == 0:
                continue
            alguma_indevida_com_dado = True
            if comp_val != consol_val:
                indev_linhas.append(
                    f"DIV {cat}: verif. {_fmt_num(comp_val)} / "
                    f"consolidado {_fmt_num(consol_val)}"
                )
                tem_divergencia = True
            else:
                indev_linhas.append(f"OK {cat}: {_fmt_num(consol_val)}")

        if not alguma_indevida_com_dado:
            indev_linhas.append("OK indevidas: todas zeradas")

        linhas.extend(indev_linhas)

        # Alerta safeframe: se > 10% do total de impressões do consolidado
        safeframe_verif = verif_indev.get("safeframe", 0)
        entregue_consol = consol_row.get("entregue", 0)
        if safeframe_verif > 0 and entregue_consol > 0:
            pct_sf = safeframe_verif / entregue_consol * 100
            if pct_sf > 10:
                linhas.append(
                    f"ALERTA safeframe: {_fmt_num(safeframe_verif)} imp "
                    f"({pct_sf:.1f}% do total) — alto índice, verificar rastreamento"
                )
    else:
        linhas.append("? indevidas: sem arquivo de verification")

    # ── Resumo de performance do consolidado (sempre exibido) ──────────────────
    impressoes_c = consol_row.get("entregue", 0)
    cliques_c    = consol_row.get("cliques", 0)
    views_c      = consol_row.get("views", 0)
    viewables_c  = consol_row.get("viewables", 0)
    va_c         = consol_row.get("viewability")

    # Campos já reportados pela comparação com comprovante (evita duplicação)
    reported = {ln.split(" ")[1].rstrip(":") for ln in linhas if " " in ln}

    if "entregue" not in reported and impressoes_c:
        linhas.append(f"OK impressoes: {_fmt_num(impressoes_c)}")
    if "cliques" not in reported and cliques_c:
        linhas.append(f"OK cliques: {_fmt_num(cliques_c)}")
    if impressoes_c and cliques_c:
        linhas.append(f"OK CTR: {cliques_c / impressoes_c * 100:.2f}%")
    for _vf in ("views_start", "views_50", "views_100"):
        _vv = consol_row.get(_vf) or 0
        if _vf not in reported and _vv:
            linhas.append(f"OK {_vf}: {_fmt_num(_vv)}")
    if "views" not in reported and views_c and not any(consol_row.get(k) for k in ("views_start", "views_50", "views_100")):
        linhas.append(f"OK views: {_fmt_num(views_c)}")
    if impressoes_c and views_c:
        linhas.append(f"OK VTR: {views_c / impressoes_c * 100:.2f}%")
    if "viewables" not in reported and viewables_c:
        linhas.append(f"OK VA: {_fmt_num(viewables_c)}")
    if "viewability" not in reported and va_c is not None:
        linhas.append(f"OK VA%: {va_c:.2f}%")

    if not linhas:
        return "OK", ["OK — sem dados para comparar"]

    # ── Delta (comp/consol) vs verification (sempre ao final) ──────────────────
    # Usa verif_result_dif (sem filtro de praça) quando disponível, pois a DIF
    # deve comparar com o total do verification, não com o subconjunto filtrado.
    verif_for_dif = verif_result_dif or verif_result
    if verif_for_dif is not None:
        base_imp    = (comp_result.get("entregue") if comp_result is not None else consol_row.get("entregue")) or 0
        verif_imp   = verif_for_dif.get("entregue_total") or verif_for_dif.get("entregue") or 0
        base_views  = (comp_result.get("views") if comp_result is not None else consol_row.get("views")) or 0
        # Alguns verifs reportam contagem equivalente em "entregue".
        verif_views = (verif_for_dif.get("views") or verif_imp or 0)
        base_label  = "comp" if comp_result is not None else "consol"

        # Regra: se houver views no consolidado, o DIF principal deve usar views.
        if consol_row.get("views"):
            if base_views and verif_views:
                delta = base_views - verif_views
                pct   = delta / base_views * 100
                linhas.append(
                    f"DIF views: {base_label} {_fmt_num(base_views)} / "
                    f"verif {_fmt_num(verif_views)} = {_fmt_num(delta)} ({pct:+.1f}%)"
                )
        else:
            if base_imp and verif_imp:
                delta = base_imp - verif_imp
                pct   = delta / base_imp * 100
                linhas.append(
                    f"DIF impressoes: {base_label} {_fmt_num(base_imp)} / "
                    f"verif {_fmt_num(verif_imp)} = {_fmt_num(delta)} ({pct:+.1f}%)"
                )
            if base_views and verif_views:
                delta = base_views - verif_views
                pct   = delta / base_views * 100
                linhas.append(
                    f"DIF views: {base_label} {_fmt_num(base_views)} / "
                    f"verif {_fmt_num(verif_views)} = {_fmt_num(delta)} ({pct:+.1f}%)"
                )

    return ("DIVERGENCIA" if tem_divergencia else "OK"), linhas


# ── Engine principal ────────────────────────────────────────────────────────────
def verificar(
    consolidado_path: str,
    adserver: str,
    comp_paths: list[str],
    verif_paths: list[str],
    data_ini: date | None = None,
    data_fim: date | None = None,
    output_path: str | None = None,
    url_sample_pct: int = 10,
    view_rules: list[dict] | None = None,
    praca: str | None = None,
) -> dict:
    """
    Verifica um consolidado contra comprovantes e arquivos de verification.

    consolidado_path — arquivo xlsx no padrão do template (29 colunas)
    adserver         — identificador do adserver (ex.: "adforce", "admotion")
    comp_paths       — arquivos de comprovante de entrega
    verif_paths      — arquivos de verification de URL (opcional)
    data_ini / data_fim — filtro de período (opcional)
    output_path      — caminho de saída; padrão: <nome> - Verificado.xlsx
    praca            — sigla do estado para filtrar verification (ex.: "SP")

    Retorna dict com:
      {
        "output":           str,
        "veiculos":         list[dict],
        "sem_comprovante":  list[str],
        "sem_consolidado":       list[str],
        "sem_consolidado_verif": list[str],
        "sem_consolidado_comp":  list[str],
        "parse_errors":     list[dict],
        "url_sample":       list[dict],
      }
    """
    if adserver not in PARSER_MAP:
        raise ValueError(
            f"Adserver '{adserver}' não reconhecido. "
            f"Opções: {', '.join(PARSER_MAP)}"
        )

    mod = importlib.import_module(PARSER_MAP[adserver])
    parse_comp  = mod.parse_comprovante
    parse_verif = mod.parse_verif

    parse_errors: list[dict] = []

    # ── Parsear verification files ─────────────────────────────────────────────
    verif_raw: list[dict] = []
    url_pool: list[dict] = []

    for vp in verif_paths:
        try:
            results = parse_verif(vp, data_ini=data_ini, data_fim=data_fim, praca=praca)
            veiculos_encontrados = [r["veiculo"] for r in results]
            print(
                f"[verif] {Path(vp).name}: {len(results)} veículos"
                f" → {veiculos_encontrados}"
                f" | filtro: {data_ini}–{data_fim}",
                file=sys.stderr,
            )
            verif_raw.extend(results)
            for r in results:
                url_pool.extend(r.pop("url_sample", []))
        except Exception as e:
            print(f"[verif] ERRO {Path(vp).name}: {e}", file=sys.stderr)
            parse_errors.append({"arquivo": Path(vp).name, "erro": str(e)})

    # ── Parsear comprovante files ──────────────────────────────────────────────
    comp_raw: list[dict] = []

    for cp in comp_paths:
        try:
            comp_raw.extend(parse_comp(cp, data_ini=data_ini, data_fim=data_fim))
        except Exception as e:
            parse_errors.append({"arquivo": Path(cp).name, "erro": str(e)})

    # Safeframe é limitação técnica, não conteúdo indevido — exclui do pool de IA.
    # Mantém apenas URLs de categorias indevidas reconhecidas pelo consolidado;
    # safeframe é limitação técnica (não conteúdo indevido) e também é excluído.
    url_pool = [
        item for item in url_pool
        if normaliza_categoria(item.get("categoria", "")) not in (None, "safeframe")
    ]

    # ── Agrupa URLs duplicadas por (url, categoria, veiculo) somando impressões ──
    # O mesmo URL pode aparecer em múltiplas linhas do verification file. Agrupar
    # evita re-análise pela IA e dá o total real de impressões por URL.
    grouped: dict[tuple[str, str, str], dict] = {}
    for item in url_pool:
        key = (item["url"], item["categoria"], item["veiculo"])
        if key in grouped:
            grouped[key]["impressoes"] = grouped[key].get("impressoes", 0) + item.get("impressoes", 0)
        else:
            grouped[key] = dict(item)
    url_pool = list(grouped.values())

    # ── Amostra de URLs: pct% por categoria indevida (0 = todas) ──────────────
    # Parsers devolvem o pool completo (reservoir ≤ 500); amostragem feita aqui.
    if url_pool:
        if url_sample_pct == 0:
            url_sample: list[dict] = list(url_pool)
        else:
            by_cat: dict[str, list] = defaultdict(list)
            for item in url_pool:
                by_cat[item["categoria"]].append(item)
            url_sample = []
            for items in by_cat.values():
                n = max(1, len(items) * url_sample_pct // 100)
                url_sample.extend(random.sample(items, min(n, len(items))))
    else:
        url_sample = []

    # ── Parseia verification SEM filtro de praça p/ DIF (que não deve filtrar por estado) ──
    verif_raw_unfiltered: list[dict] = []
    for vp in verif_paths:
        try:
            results = parse_verif(vp, data_ini=data_ini, data_fim=data_fim, praca=None)
            verif_raw_unfiltered.extend(results)
        except Exception as e:
            print(f"[verif/unfiltered] ERRO {Path(vp).name}: {e}", file=sys.stderr)
    verif_norm_unfiltered = _merge_by_veiculo(verif_raw_unfiltered)
    verif_names_unfiltered = list(verif_norm_unfiltered.keys())

    # ── Índices por veículo normalizado (somando múltiplos arquivos) ──────────
    verif_norm  = _merge_by_veiculo(verif_raw)
    comp_norm   = _merge_by_veiculo(comp_raw)
    verif_names = list(verif_norm.keys())
    comp_names  = list(comp_norm.keys())

    # ── Mapa de regras de visualização: (veiculo_norm, duracao | None) → criterio
    rule_map: dict[tuple, str] = {}
    for r in (view_rules or []):
        dur = int(r["secundagem"]) if r.get("secundagem") else None
        rule_map[(_normalize(r["veiculo"]), dur)] = r["criterio"]

    # ── Aplica regras de visualização: reconstrói campos de views em comp_norm ──
    # comp_raw tem entradas separadas por (veiculo, duracao); comp_norm as fundiu.
    # Para veículos com regras, zera os campos de views e reconstrói somando apenas
    # o campo correto de cada entrada por duração.
    if rule_map:
        _vf_map = {"start": "views_start", "50": "views_50", "100": "views_100"}
        _view_fields = set(_vf_map.values()) | {"views"}
        vehicles_with_rules = {vk for (vk, _) in rule_map}
        for vk in vehicles_with_rules:
            if vk in comp_norm:
                for f in _view_fields:
                    comp_norm[vk][f] = None
        for r in comp_raw:
            vk = _normalize(r["veiculo"])
            if vk not in vehicles_with_rules or vk not in comp_norm:
                continue
            dur = r.get("duracao_segundos")
            rule = rule_map.get((vk, dur)) or rule_map.get((vk, None))
            if rule is None:
                continue
            field = _vf_map[rule]
            val = r.get(field) or 0
            comp_norm[vk][field] = (comp_norm[vk].get(field) or 0) + val

    # ── Ler consolidado ───────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(consolidado_path)
    ws = wb.active
    try:
        consol_rows, col_devolutiva_bi = _read_consolidado(ws)

        # ── Match fuzzy veículo-a-veículo ─────────────────────────────────────
        resultado_veiculos: list[dict] = []
        matched_verif_names: set[str] = set()
        matched_comp_names:  set[str] = set()

        for crow in consol_rows:
            consol_norm = _normalize(crow["veiculo"])

            verif_match, verif_score = _fuzzy_match(consol_norm, verif_names, verif_norm)
            comp_match,  comp_score  = _fuzzy_match(consol_norm, comp_names,  comp_norm)
            verif_match_dif, _       = _fuzzy_match(consol_norm, verif_names_unfiltered, verif_norm_unfiltered)

            if verif_match:
                matched_verif_names.add(_normalize(verif_match["veiculo"]))
            if comp_match:
                matched_comp_names.add(_normalize(comp_match["veiculo"]))

            if verif_match is None and comp_match is None:
                devolutiva = "PENDENTE: comprovante/verification nao localizado"
                status = "PENDENTE"
                match_name = None
                score = 0
            else:
                status, linhas = _compare(crow, comp_match, verif_match, verif_match_dif)
                devolutiva = "\n".join(linhas)
                match_name = (verif_match or comp_match or {}).get("veiculo")
                score = verif_score or comp_score

            resultado_veiculos.append({
                "veiculo":        crow["veiculo"],
                "status":         status,
                "devolutiva":     devolutiva,
                "match":          match_name,
                "score":          score,
                "formato":        (verif_match or {}).get("formato_detectado"),
                "entregue_consol": crow["entregue"],
            })

            cell = ws.cell(row=crow["row_idx"], column=col_devolutiva_bi)
            cell.value = devolutiva

            # Aplica cor de fundo baseada no status/conteúdo da devolutiva
            if status == "DIVERGENCIA":
                cell.fill = COLOR_DIV
            elif "ALERTA" in devolutiva:
                cell.fill = COLOR_ALERTA
            elif status == "PENDENTE":
                cell.fill = COLOR_PENDENTE
            else:
                cell.fill = COLOR_OK

        # ── Veículos sem entrada no consolidado ──────────────────────────────
        sem_consolidado_verif: list[str] = []
        sem_consolidado_comp: list[str] = []
        seen: set[str] = set()
        for norm, r in verif_norm.items():
            if norm not in matched_verif_names:
                name = r["veiculo"]
                if name not in seen:
                    sem_consolidado_verif.append(name)
                    seen.add(name)
        for norm, r in comp_norm.items():
            if norm not in matched_comp_names:
                name = r["veiculo"]
                if name not in seen:
                    sem_consolidado_comp.append(name)
                    seen.add(name)

        # ── Salvar arquivo verificado ─────────────────────────────────────────
        if output_path is None:
            p = Path(consolidado_path)
            output_path = str(p.parent / (p.stem + " - Verificado" + p.suffix))

        wb.save(output_path)
    finally:
        wb.close()

    return {
        "output":           output_path,
        "veiculos":         resultado_veiculos,
        "sem_comprovante":  [r["veiculo"] for r in resultado_veiculos if r["status"] == "PENDENTE"],
        "sem_consolidado":       sem_consolidado_verif + sem_consolidado_comp,
        "sem_consolidado_verif": sem_consolidado_verif,
        "sem_consolidado_comp":  sem_consolidado_comp,
        "parse_errors":     parse_errors,
        "url_sample":       url_sample,
    }


# ── CLI ─────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Engine de verificação de consolidados SECOM")
    ap.add_argument("consolidado",  help="Consolidado .xlsx (template 29 colunas)")
    ap.add_argument("--adserver", required=True,
                    choices=list(PARSER_MAP.keys()),
                    help="Adserver dos arquivos (ex.: adforce, admotion)")
    ap.add_argument("--comp", nargs="+", default=[],
                    metavar="ARQUIVO", help="Arquivo(s) de comprovante de entrega")
    ap.add_argument("--verif", nargs="*", default=[],
                    metavar="ARQUIVO", help="Arquivo(s) de verification de URL (opcional)")
    ap.add_argument("--ini", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--fim", default=None, metavar="DD/MM/YYYY")
    ap.add_argument("--output", default=None, metavar="PATH",
                    help="Caminho de saída (padrão: <consolidado> - Verificado.xlsx)")
    ap.add_argument("--url-pct", type=int, default=10, metavar="PCT",
                    help="% de URLs indevidas a analisar via IA (0 = todas)")
    ap.add_argument("--view-rules", default=None,
                    help="JSON array de regras de visualização por veículo")
    ap.add_argument("--praca", default=None, metavar="UF",
                    help="Sigla do estado para filtrar verification (ex.: SP)")
    args = ap.parse_args()

    from parser_utils import cli_date  # noqa: E402

    try:
        resultado = verificar(
            args.consolidado,
            args.adserver,
            args.comp,
            args.verif or [],
            data_ini=cli_date(args.ini),
            data_fim=cli_date(args.fim),
            output_path=args.output,
            url_sample_pct=args.url_pct,
            view_rules=json.loads(args.view_rules) if args.view_rules else None,
            praca=args.praca,
        )
        # Resumo legível → stderr (não polui o JSON que o Node.js consome)
        print(f"\nArquivo gerado: {resultado['output']}", file=sys.stderr)
        print(f"\n{'VEÍCULO':<35} {'STATUS':<12} {'MATCH':<35} {'SCORE':>5}", file=sys.stderr)
        print("-" * 90, file=sys.stderr)
        for v in resultado["veiculos"]:
            print(f"{v['veiculo']:<35} {v['status']:<12} {str(v.get('match') or ''):<35} {v.get('score',0):>5.0f}", file=sys.stderr)
        if resultado["sem_consolidado_verif"]:
            print(f"\nVerification sem consolidado ({len(resultado['sem_consolidado_verif'])}):", file=sys.stderr)
            for n in resultado["sem_consolidado_verif"]:
                print(f"  • {n}", file=sys.stderr)
        if resultado["sem_consolidado_comp"]:
            print(f"\nComprovante sem consolidado ({len(resultado['sem_consolidado_comp'])}):", file=sys.stderr)
            for n in resultado["sem_consolidado_comp"]:
                print(f"  - {n}", file=sys.stderr)
        if resultado["parse_errors"]:
            print(f"\nErros de parse ({len(resultado['parse_errors'])}):", file=sys.stderr)
            for e in resultado["parse_errors"]:
                print(f"  - {e['arquivo']}: {e['erro']}", file=sys.stderr)
        # JSON puro → stdout (consumido pelo Node.js)
        print(json.dumps(resultado, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"erro": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)
